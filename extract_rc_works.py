"""
Fixed Extraction script for RC Works sheet
Uses actual Excel codes, includes sheet name in cell references, and detects bold subcategories
"""

from extractor_base import BaseExtractor
import pandas as pd
import re
from openpyxl import load_workbook

class RCWorksExtractor(BaseExtractor):
    def __init__(self, excel_file='MJD-PRICELIST.xlsx'):
        super().__init__(excel_file, 'RC works')
        self.workbook = None
        self.worksheet = None
        self.current_subcategory = 'RC Works'  # Default subcategory
        
    def load_workbook_for_formatting(self):
        """Load workbook with openpyxl to access formatting"""
        try:
            self.workbook = load_workbook(self.excel_file, data_only=True)
            self.worksheet = self.workbook[self.sheet_name]
            print(f"Loaded workbook for formatting detection")
        except Exception as e:
            print(f"Warning: Could not load workbook for formatting: {e}")
            self.worksheet = None
    
    def is_row_bold(self, row_idx):
        """Check if all non-empty cells in a row are bold"""
        if not self.worksheet:
            return False
        
        try:
            # Excel rows are 1-indexed
            excel_row = row_idx + 1
            row_is_bold = False
            has_content = False
            
            # Check first 5 columns for bold text
            for col in range(1, 6):
                cell = self.worksheet.cell(row=excel_row, column=col)
                if cell.value:
                    has_content = True
                    if cell.font and cell.font.bold:
                        row_is_bold = True
                    else:
                        # If any cell with content is not bold, row is not fully bold
                        return False
            
            return has_content and row_is_bold
        except Exception as e:
            return False
    
    def extract_description(self, row, start_col=1):
        """Extract and clean description for RC works"""
        description_parts = []
        
        # Column B (index 1) is usually the main description
        if len(row) > 1 and pd.notna(row[1]):
            desc = str(row[1]).strip()
            if desc and not self.is_unit(desc):
                description_parts.append(desc)
        
        # Column C (index 2) might have continuation or additional info
        if len(row) > 2 and pd.notna(row[2]):
            part = str(row[2]).strip()
            # Only add if it's not a unit and not a number
            if part and not self.is_unit(part) and not re.match(r'^[\d,\.]+$', part):
                description_parts.append(part)
        
        description = ' '.join(description_parts)
        
        # Clean and expand RC-specific abbreviations
        replacements = {
            ' rc ': ' reinforced concrete ',
            ' r.c. ': ' reinforced concrete ',
            ' conc ': ' concrete ',
            ' reinf ': ' reinforcement ',
            ' fwk ': ' formwork ',
            ' ne ': ' not exceeding ',
            ' n.e. ': ' not exceeding ',
            ' thk ': ' thick ',
            ' dia ': ' diameter ',
            ' c/c ': ' centers ',
            ' bwys ': ' both ways ',
            ' ew ': ' each way ',
            ' t&b ': ' top and bottom ',
            ' u/s ': ' underside ',
            ' o/a ': ' overall ',
            ' incl ': ' including ',
            ' excl ': ' excluding ',
            ' horiz ': ' horizontal ',
            ' vert ': ' vertical ',
        }
        
        for old, new in replacements.items():
            description = description.replace(old, new)
        
        # Fix patterns
        description = re.sub(r'(\d+)thk', r'\1mm thick', description)
        description = re.sub(r'(\d+)dia', r'\1mm diameter', description)
        
        # Clean up spaces
        description = ' '.join(description.split())
        
        return description
    
    def extract_unit(self, row):
        """Extract unit from row - primarily column E (index 4)"""
        # Check column E first (index 4) - this is the primary unit column
        if len(row) > 4 and pd.notna(row[4]):
            value = str(row[4]).strip()
            if self.is_unit(value):
                return self.standardize_unit(value)
        
        # Then check columns C and D as fallback
        for col_idx in [2, 3]:
            if col_idx < len(row) and pd.notna(row[col_idx]):
                value = str(row[col_idx]).strip()
                if self.is_unit(value):
                    return self.standardize_unit(value)
        
        # Infer from description if not found
        desc = self.extract_description(row)
        desc_lower = desc.lower()
        
        # RC works specific patterns
        if 'reinforcement' in desc_lower or 'rebar' in desc_lower or 'steel' in desc_lower:
            if 'mesh' in desc_lower:
                return 'm2'
            return 'kg'
        elif 'concrete' in desc_lower:
            if any(word in desc_lower for word in ['slab', 'surface', 'topping', 'screed', 'blinding']):
                if 'thick' in desc_lower:
                    return 'm2'
            elif any(word in desc_lower for word in ['beam', 'column', 'wall', 'foundation']):
                return 'm3'
            return 'm3'
        elif 'formwork' in desc_lower or 'shutter' in desc_lower:
            if 'edge' in desc_lower or 'linear' in desc_lower:
                return 'm'
            return 'm2'
        elif 'mesh' in desc_lower:
            return 'm2'
        elif any(word in desc_lower for word in ['joint', 'groove', 'chase']):
            return 'm'
        
        return 'item'
    
    def determine_subcategory(self, description):
        """Determine subcategory based on description for RC works"""
        desc_lower = description.lower()
        
        # RC works subcategories
        if 'concrete' in desc_lower:
            if 'foundation' in desc_lower:
                return 'Concrete Foundations'
            elif 'slab' in desc_lower:
                if 'ground' in desc_lower:
                    return 'Ground Slabs'
                elif 'suspend' in desc_lower:
                    return 'Suspended Slabs'
                else:
                    return 'Concrete Slabs'
            elif 'beam' in desc_lower:
                return 'Concrete Beams'
            elif 'column' in desc_lower:
                return 'Concrete Columns'
            elif 'wall' in desc_lower:
                if 'retaining' in desc_lower:
                    return 'Retaining Walls'
                else:
                    return 'Concrete Walls'
            elif 'stair' in desc_lower:
                return 'Concrete Stairs'
            elif 'blinding' in desc_lower:
                return 'Blinding'
            else:
                return 'Concrete Works'
        elif 'reinforcement' in desc_lower or 'rebar' in desc_lower or 'steel' in desc_lower:
            if 'mesh' in desc_lower:
                return 'Mesh Reinforcement'
            elif 'bar' in desc_lower:
                return 'Bar Reinforcement'
            else:
                return 'Steel Reinforcement'
        elif 'formwork' in desc_lower or 'shutter' in desc_lower:
            if 'slab' in desc_lower or 'soffit' in desc_lower:
                return 'Slab Formwork'
            elif 'wall' in desc_lower or 'vertical' in desc_lower:
                return 'Wall Formwork'
            elif 'beam' in desc_lower:
                return 'Beam Formwork'
            elif 'column' in desc_lower:
                return 'Column Formwork'
            else:
                return 'Formwork'
        elif 'waterproof' in desc_lower:
            return 'Waterproofing'
        elif 'joint' in desc_lower:
            return 'Joints and Accessories'
        elif 'finish' in desc_lower:
            return 'Concrete Finishes'
        else:
            return 'RC Works'
    
    def generate_keywords(self, description):
        """Generate search keywords for RC works"""
        keywords = []
        desc_lower = description.lower()
        
        # Extract concrete grades
        grades = re.findall(r'c\d+(?:/\d+)?', desc_lower)
        keywords.extend(grades[:2])
        
        # Extract thickness
        thickness = re.findall(r'\d+(?:mm)?\s*thick', desc_lower)
        keywords.extend([t.replace(' ', '') for t in thickness[:1]])
        
        # Key RC terms
        terms = ['concrete', 'reinforcement', 'formwork', 'rebar', 'mesh',
                 'slab', 'beam', 'column', 'wall', 'foundation']
        
        for term in terms:
            if term in desc_lower:
                keywords.append(term)
        
        return keywords[:5]
    
    def extract_items(self):
        """Main extraction method with bold subcategory detection"""
        if self.df is None:
            self.load_sheet()
        
        # Load workbook for formatting detection
        self.load_workbook_for_formatting()
        
        print(f"\nExtracting items from {self.sheet_name}...")
        print(f"Starting from row 12...")
        
        items = []
        current_subcategory = 'RC Works'  # Default subcategory
        rows_processed = 0
        rows_skipped = 0
        
        # Process all rows starting from row 12 (where data begins)
        start_row = 11  # Row 12 in Excel (0-indexed)
        
        for row_idx in range(start_row, len(self.df)):
            row = self.df.iloc[row_idx]
            
            # Skip if row is completely empty (but allow single-cell rows for potential headers)
            if row.notna().sum() < 1:
                continue
            
            # Check if this row is bold (potential subcategory header)
            if self.is_row_bold(row_idx):
                # Extract text from the row to use as subcategory
                subcategory_text = ''
                for col_idx in range(min(5, len(row))):
                    if pd.notna(row[col_idx]):
                        text = str(row[col_idx]).strip()
                        if text and not self.is_unit(text):
                            subcategory_text = text
                            break
                
                if subcategory_text:
                    current_subcategory = subcategory_text
                    print(f"Found subcategory at row {row_idx + 1}: {current_subcategory}")
                    continue  # Skip this row as it's a header
            
            # Check if this is a data row
            # Simple check - if column A has something and column B has text
            first_col = row[0] if 0 < len(row) else None
            second_col = row[1] if len(row) > 1 else None
            
            # Skip if column A is empty
            if pd.isna(first_col):
                continue
                
            # Skip if column B is empty or too short
            if pd.isna(second_col):
                continue
                
            first_str = str(first_col).strip()
            second_str = str(second_col).strip()
            
            # Skip if column A has nothing meaningful
            if not first_str or first_str.lower() in ['', 'nan', 'none']:
                continue
                
            # Skip if column B is too short to be a description
            if len(second_str) < 5:
                continue
                
            # Skip if column B is just a unit
            if self.is_unit(second_str):
                continue
            
            rows_processed += 1
            
            # Extract code (actual Excel value)
            code = self.extract_code(row)
            
            # Extract description
            description = self.extract_description(row)
            
            # Skip if no valid description
            if not description or len(description) < 5:
                rows_skipped += 1
                continue
            
            # Extract unit
            unit = self.extract_unit(row)
            
            # Extract rate and column index
            rate, rate_col_idx = self.extract_rate(row)
            
            # Use current subcategory (from bold header) or determine from keywords as fallback
            if current_subcategory and current_subcategory != 'RC Works':
                subcategory = current_subcategory
            else:
                subcategory = self.determine_subcategory(description)
            
            # Generate keywords
            keywords = self.generate_keywords(description)
            
            # Create item with actual code
            item = self.create_item(
                row_idx=row_idx,
                row=row,
                code=code,
                description=description,
                unit=unit,
                subcategory=subcategory,
                rate=rate,
                rate_col_idx=rate_col_idx,
                keywords=keywords
            )
            
            items.append(item)
        
        self.extracted_items = items
        print(f"Processed {rows_processed} rows, skipped {rows_skipped} due to short descriptions")
        print(f"Extracted {len(items)} valid items from {self.sheet_name}")
        return items

def main():
    print("="*60)
    print("RC WORKS SHEET EXTRACTION (FIXED)")
    print("="*60)
    
    extractor = RCWorksExtractor()
    items = extractor.extract_items()
    
    if items:
        # Show sample
        print("\nSample extracted items:")
        for item in items[:3]:
            print(f"\nID: {item['id']}")
            print(f"  Code: {item['code']}")
            print(f"  Description: {item['description'][:60]}...")
            print(f"  Unit: {item['unit']}")
            print(f"  Rate: {item['rate']}")
            print(f"  Cell Ref: {item['cellRate_reference']}")
        
        extractor.save_output()
        
        print(f"\nTotal items: {len(items)}")
        print(f"Items with rates: {sum(1 for i in items if i['rate'])}")
        print(f"Items with cell refs: {sum(1 for i in items if i['cellRate_reference'])}")

if __name__ == "__main__":
    main()