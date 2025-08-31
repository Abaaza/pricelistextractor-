"""
MJD Pricelist Extraction Script
Extracts pricing data from all sheets with custom logic per sheet structure
"""

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import numpy as np
import json
import csv
import re
from typing import Dict, List, Optional, Tuple, Any
from dataclasses import dataclass, asdict
from pathlib import Path
import openai
import os
from datetime import datetime
import hashlib

# OpenAI Configuration
OPENAI_API_KEY = os.getenv('OPENAI_API_KEY', '')  # Set your API key as environment variable

@dataclass
class CellRate:
    """Cell reference for rate updates"""
    reference: str
    sheetName: str
    rate: float

@dataclass
class PriceItem:
    """Schema for price list item"""
    id: str
    code: Optional[str] = None
    ref: Optional[str] = None
    description: Optional[str] = None
    keywords: Optional[List[str]] = None
    unit: Optional[str] = None
    category: Optional[str] = None
    subCategory: Optional[str] = None
    cellRates: Optional[Dict[str, CellRate]] = None
    patterns: Optional[List[Dict]] = None

class SheetExtractor:
    """Base class for sheet-specific extraction logic"""
    
    def __init__(self, sheet_name: str, worksheet):
        self.sheet_name = sheet_name
        self.worksheet = worksheet
        self.items = []
        
    def get_cell_reference(self, row: int, col: int) -> str:
        """Get Excel-style cell reference (e.g., A1, B2)"""
        return f"{get_column_letter(col)}{row}"
    
    def extract_header_info(self) -> Dict:
        """Extract header and subcategory information"""
        # Default implementation - can be overridden
        headers = {}
        for row in range(1, min(20, self.worksheet.max_row + 1)):
            for col in range(1, min(10, self.worksheet.max_column + 1)):
                cell_value = self.worksheet.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str):
                    if any(keyword in str(cell_value).lower() for keyword in ['schedule', 'works', 'section']):
                        headers['main_header'] = cell_value
                        headers['header_row'] = row
                        break
            if headers:
                break
        return headers
    
    def clean_value(self, value: Any) -> Any:
        """Clean cell values"""
        if value is None:
            return None
        if isinstance(value, str):
            return value.strip()
        return value
    
    def is_valid_rate(self, value: Any) -> bool:
        """Check if value is a valid rate"""
        if value is None:
            return False
        try:
            rate = float(value)
            return 0 < rate < 1000000  # Reasonable rate range
        except:
            return False
    
    def extract_items(self) -> List[PriceItem]:
        """Main extraction method - must be implemented by subclasses"""
        raise NotImplementedError("Each sheet extractor must implement extract_items")

class GroundworksExtractor(SheetExtractor):
    """Extractor for Groundworks sheet"""
    
    def extract_items(self) -> List[PriceItem]:
        items = []
        current_subcategory = None
        item_counter = 1
        
        # Find data start row (usually after headers)
        data_start_row = 10
        for row in range(1, min(30, self.worksheet.max_row + 1)):
            cell = self.worksheet.cell(row=row, column=2)
            if cell.value and 'Description' in str(cell.value):
                data_start_row = row + 2
                break
        
        # Process rows
        for row in range(data_start_row, self.worksheet.max_row + 1):
            # Check for subcategory headers (bold or merged cells)
            cell_b = self.worksheet.cell(row=row, column=2)
            
            # Check if this is a subcategory header
            if cell_b.value and cell_b.font and cell_b.font.bold:
                current_subcategory = self.clean_value(cell_b.value)
                continue
            
            # Extract item data
            description = self.clean_value(cell_b.value)
            if not description or description == '':
                continue
                
            # Skip non-item rows
            if any(skip in str(description).lower() for skip in ['total', 'subtotal', 'carried forward', 'brought forward']):
                continue
            
            # Get quantity, unit, rate columns (typically columns D, E, F)
            quantity = self.worksheet.cell(row=row, column=4).value
            unit = self.clean_value(self.worksheet.cell(row=row, column=5).value)
            rate = self.worksheet.cell(row=row, column=6).value
            
            # Only process rows with valid data
            if description and (rate or quantity):
                # Generate ID
                item_id = f"GW_{item_counter:04d}_{self.sheet_name}"
                
                # Get cell references for rates
                cell_rates = {}
                if self.is_valid_rate(rate):
                    cell_rates['cellRate1'] = CellRate(
                        reference=self.get_cell_reference(row, 6),
                        sheetName=self.sheet_name,
                        rate=float(rate) if rate else 0.0
                    )
                
                # Check for additional rate columns
                for col_offset, rate_name in enumerate(['cellRate2', 'cellRate3', 'cellRate4'], start=1):
                    additional_rate = self.worksheet.cell(row=row, column=6 + col_offset).value
                    if self.is_valid_rate(additional_rate):
                        cell_rates[rate_name] = CellRate(
                            reference=self.get_cell_reference(row, 6 + col_offset),
                            sheetName=self.sheet_name,
                            rate=float(additional_rate)
                        )
                
                item = PriceItem(
                    id=item_id,
                    code=f"GW{item_counter:04d}",
                    description=description,
                    unit=unit,
                    category=self.sheet_name,
                    subCategory=current_subcategory,
                    cellRates=cell_rates if cell_rates else None
                )
                
                items.append(item)
                item_counter += 1
        
        return items

class RCWorksExtractor(SheetExtractor):
    """Extractor for RC Works sheet"""
    
    def extract_items(self) -> List[PriceItem]:
        items = []
        current_subcategory = None
        item_counter = 1
        
        # RC Works specific structure
        data_start_row = 10
        for row in range(1, min(30, self.worksheet.max_row + 1)):
            cell = self.worksheet.cell(row=row, column=2)
            if cell.value and 'Description' in str(cell.value):
                data_start_row = row + 2
                break
        
        for row in range(data_start_row, self.worksheet.max_row + 1):
            # Column B for descriptions
            cell_b = self.worksheet.cell(row=row, column=2)
            
            # Check for subcategory headers
            if cell_b.value:
                value_str = str(cell_b.value)
                # Common RC subcategories
                if any(cat in value_str for cat in ['In-situ Concrete', 'Formwork', 'Reinforcement', 
                                                     'Precast', 'Post-tensioning', 'Sundries']):
                    current_subcategory = self.clean_value(cell_b.value)
                    continue
            
            description = self.clean_value(cell_b.value)
            if not description:
                continue
                
            # Skip summary rows
            if any(skip in str(description).lower() for skip in ['total', 'subtotal', 'carried']):
                continue
            
            # Get data from standard columns
            quantity = self.worksheet.cell(row=row, column=4).value
            unit = self.clean_value(self.worksheet.cell(row=row, column=5).value)
            rate = self.worksheet.cell(row=row, column=6).value
            
            if description and (rate or quantity):
                item_id = f"RC_{item_counter:04d}_{self.sheet_name}"
                
                cell_rates = {}
                if self.is_valid_rate(rate):
                    cell_rates['cellRate1'] = CellRate(
                        reference=self.get_cell_reference(row, 6),
                        sheetName=self.sheet_name,
                        rate=float(rate) if rate else 0.0
                    )
                
                item = PriceItem(
                    id=item_id,
                    code=f"RC{item_counter:04d}",
                    description=description,
                    unit=unit,
                    category=self.sheet_name,
                    subCategory=current_subcategory,
                    cellRates=cell_rates if cell_rates else None
                )
                
                items.append(item)
                item_counter += 1
        
        return items

class DrainageExtractor(SheetExtractor):
    """Extractor for Drainage sheet"""
    
    def extract_items(self) -> List[PriceItem]:
        items = []
        current_subcategory = None
        item_counter = 1
        
        # Drainage has different structure - find actual data start
        data_start_row = 15
        for row in range(1, min(50, self.worksheet.max_row + 1)):
            # Look for typical drainage items
            cell = self.worksheet.cell(row=row, column=2)
            if cell.value and any(term in str(cell.value).lower() for term in ['pipe', 'excavat', 'manhole']):
                data_start_row = row
                break
        
        for row in range(data_start_row, self.worksheet.max_row + 1):
            # Get description from column B or C depending on sheet structure
            description = None
            for col in [2, 3]:
                cell_value = self.worksheet.cell(row=row, column=col).value
                if cell_value and len(str(cell_value)) > 5:
                    description = self.clean_value(cell_value)
                    break
            
            if not description:
                continue
            
            # Check for subcategory markers
            desc_lower = description.lower()
            if any(cat in desc_lower for cat in ['below ground drainage', 'above ground', 
                                                  'manholes', 'gullies', 'pipework']):
                current_subcategory = description
                continue
            
            # Skip non-items
            if any(skip in desc_lower for skip in ['total', 'subtotal', 'labour only']):
                continue
            
            # Find rate columns (may vary in position)
            rate = None
            rate_col = None
            for col in range(5, min(15, self.worksheet.max_column + 1)):
                cell_value = self.worksheet.cell(row=row, column=col).value
                if self.is_valid_rate(cell_value):
                    rate = cell_value
                    rate_col = col
                    break
            
            # Get unit (usually before rate column)
            unit = None
            if rate_col:
                unit = self.clean_value(self.worksheet.cell(row=row, column=rate_col - 1).value)
            
            if description and rate:
                item_id = f"DR_{item_counter:04d}_{self.sheet_name}"
                
                cell_rates = {}
                if rate and rate_col:
                    cell_rates['cellRate1'] = CellRate(
                        reference=self.get_cell_reference(row, rate_col),
                        sheetName=self.sheet_name,
                        rate=float(rate)
                    )
                
                item = PriceItem(
                    id=item_id,
                    code=f"DR{item_counter:04d}",
                    description=description,
                    unit=unit,
                    category=self.sheet_name,
                    subCategory=current_subcategory,
                    cellRates=cell_rates if cell_rates else None
                )
                
                items.append(item)
                item_counter += 1
        
        return items

class ServicesExtractor(SheetExtractor):
    """Extractor for Services sheet"""
    
    def extract_items(self) -> List[PriceItem]:
        items = []
        current_subcategory = None
        item_counter = 1
        
        # Services typically includes M&E items
        data_start_row = 10
        
        for row in range(data_start_row, self.worksheet.max_row + 1):
            description = self.clean_value(self.worksheet.cell(row=row, column=2).value)
            
            if not description:
                continue
            
            # Check for service subcategories
            if any(cat in str(description).lower() for cat in ['electrical', 'mechanical', 'plumbing', 
                                                               'hvac', 'fire alarm', 'data']):
                current_subcategory = description
                continue
            
            # Standard column positions for services
            quantity = self.worksheet.cell(row=row, column=4).value
            unit = self.clean_value(self.worksheet.cell(row=row, column=5).value)
            rate = self.worksheet.cell(row=row, column=6).value
            
            if description and self.is_valid_rate(rate):
                item_id = f"SV_{item_counter:04d}_{self.sheet_name}"
                
                cell_rates = {
                    'cellRate1': CellRate(
                        reference=self.get_cell_reference(row, 6),
                        sheetName=self.sheet_name,
                        rate=float(rate)
                    )
                }
                
                item = PriceItem(
                    id=item_id,
                    code=f"SV{item_counter:04d}",
                    description=description,
                    unit=unit,
                    category=self.sheet_name,
                    subCategory=current_subcategory,
                    cellRates=cell_rates
                )
                
                items.append(item)
                item_counter += 1
        
        return items

class ExternalWorksExtractor(SheetExtractor):
    """Extractor for External Works sheet"""
    
    def extract_items(self) -> List[PriceItem]:
        items = []
        current_subcategory = None
        item_counter = 1
        
        data_start_row = 10
        
        for row in range(data_start_row, self.worksheet.max_row + 1):
            description = self.clean_value(self.worksheet.cell(row=row, column=2).value)
            
            if not description:
                continue
            
            # External works subcategories
            if any(cat in str(description).lower() for cat in ['paving', 'landscaping', 'fencing', 
                                                               'kerbs', 'roads', 'car park']):
                current_subcategory = description
                continue
            
            quantity = self.worksheet.cell(row=row, column=4).value
            unit = self.clean_value(self.worksheet.cell(row=row, column=5).value)
            rate = self.worksheet.cell(row=row, column=6).value
            
            if description and (rate or quantity):
                item_id = f"EW_{item_counter:04d}_{self.sheet_name}"
                
                cell_rates = {}
                if self.is_valid_rate(rate):
                    cell_rates['cellRate1'] = CellRate(
                        reference=self.get_cell_reference(row, 6),
                        sheetName=self.sheet_name,
                        rate=float(rate) if rate else 0.0
                    )
                
                item = PriceItem(
                    id=item_id,
                    code=f"EW{item_counter:04d}",
                    description=description,
                    unit=unit,
                    category=self.sheet_name,
                    subCategory=current_subcategory,
                    cellRates=cell_rates if cell_rates else None
                )
                
                items.append(item)
                item_counter += 1
        
        return items

class PrelimsExtractor(SheetExtractor):
    """Extractor for Preliminaries sheets"""
    
    def extract_items(self) -> List[PriceItem]:
        items = []
        current_subcategory = None
        item_counter = 1
        
        # Prelims have specific structure with weekly/monthly rates
        data_start_row = 10
        
        for row in range(data_start_row, self.worksheet.max_row + 1):
            description = self.clean_value(self.worksheet.cell(row=row, column=2).value)
            
            if not description:
                continue
            
            # Prelims subcategories
            if any(cat in str(description).lower() for cat in ['management', 'site establishment', 
                                                               'temporary works', 'welfare', 'plant']):
                current_subcategory = description
                continue
            
            # Prelims often have duration-based pricing
            duration = self.worksheet.cell(row=row, column=4).value
            unit = self.clean_value(self.worksheet.cell(row=row, column=5).value)
            if not unit and duration:
                unit = 'week' if 'week' in str(description).lower() else 'month'
            
            rate = self.worksheet.cell(row=row, column=6).value
            
            if description and (rate or duration):
                item_id = f"PR_{item_counter:04d}_{self.sheet_name}"
                
                cell_rates = {}
                if self.is_valid_rate(rate):
                    cell_rates['cellRate1'] = CellRate(
                        reference=self.get_cell_reference(row, 6),
                        sheetName=self.sheet_name,
                        rate=float(rate) if rate else 0.0
                    )
                
                item = PriceItem(
                    id=item_id,
                    code=f"PR{item_counter:04d}",
                    description=description,
                    unit=unit,
                    category=self.sheet_name,
                    subCategory=current_subcategory,
                    cellRates=cell_rates if cell_rates else None
                )
                
                items.append(item)
                item_counter += 1
        
        return items

class GenericExtractor(SheetExtractor):
    """Generic extractor for sheets without specific logic"""
    
    def extract_items(self) -> List[PriceItem]:
        items = []
        current_subcategory = None
        item_counter = 1
        
        # Try to find description column
        desc_col = 2  # Default to column B
        rate_col = 6  # Default to column F
        unit_col = 5  # Default to column E
        
        # Auto-detect column positions based on headers
        for row in range(1, min(20, self.worksheet.max_row + 1)):
            for col in range(1, min(15, self.worksheet.max_column + 1)):
                cell_value = str(self.worksheet.cell(row=row, column=col).value or '').lower()
                if 'description' in cell_value:
                    desc_col = col
                elif 'rate' in cell_value and 'rate' not in locals():
                    rate_col = col
                elif 'unit' in cell_value:
                    unit_col = col
        
        # Start extraction
        data_start_row = 10
        
        for row in range(data_start_row, self.worksheet.max_row + 1):
            description = self.clean_value(self.worksheet.cell(row=row, column=desc_col).value)
            
            if not description:
                continue
            
            # Skip totals and subtotals
            if any(skip in str(description).lower() for skip in ['total', 'subtotal', 'summary']):
                continue
            
            unit = self.clean_value(self.worksheet.cell(row=row, column=unit_col).value)
            rate = self.worksheet.cell(row=row, column=rate_col).value
            
            if description and self.is_valid_rate(rate):
                # Generate sheet-specific prefix
                prefix = ''.join([c for c in self.sheet_name[:2].upper() if c.isalpha()])
                if not prefix:
                    prefix = 'GN'
                
                item_id = f"{prefix}_{item_counter:04d}_{self.sheet_name}"
                
                cell_rates = {
                    'cellRate1': CellRate(
                        reference=self.get_cell_reference(row, rate_col),
                        sheetName=self.sheet_name,
                        rate=float(rate)
                    )
                }
                
                item = PriceItem(
                    id=item_id,
                    code=f"{prefix}{item_counter:04d}",
                    description=description,
                    unit=unit,
                    category=self.sheet_name,
                    subCategory=current_subcategory,
                    cellRates=cell_rates
                )
                
                items.append(item)
                item_counter += 1
        
        return items

class PricelistExtractor:
    """Main extractor coordinating all sheet extractors"""
    
    # Map sheet names to specific extractors
    SHEET_EXTRACTORS = {
        'Groundworks': GroundworksExtractor,
        'RC works': RCWorksExtractor,
        'Drainage': DrainageExtractor,
        'Services': ServicesExtractor,
        'External Works': ExternalWorksExtractor,
        'Prelims (full)': PrelimsExtractor,
        'Prelims (consoltd)': PrelimsExtractor,
        'Prelims Principal': PrelimsExtractor,
        # Add more specific extractors as needed
    }
    
    # Sheets to skip
    SKIP_SHEETS = ['Summary', 'Set factors & prices', 'Tender Summary', 'Budget Costings']
    
    def __init__(self, file_path: str, use_openai: bool = False):
        self.file_path = file_path
        self.use_openai = use_openai
        self.workbook = None
        self.all_items = []
        
        if use_openai and OPENAI_API_KEY:
            openai.api_key = OPENAI_API_KEY
    
    def extract_all_sheets(self) -> List[PriceItem]:
        """Extract items from all sheets"""
        self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
        
        for sheet_name in self.workbook.sheetnames:
            if sheet_name in self.SKIP_SHEETS:
                print(f"Skipping sheet: {sheet_name}")
                continue
            
            print(f"Processing sheet: {sheet_name}")
            worksheet = self.workbook[sheet_name]
            
            # Select appropriate extractor
            extractor_class = self.SHEET_EXTRACTORS.get(sheet_name, GenericExtractor)
            extractor = extractor_class(sheet_name, worksheet)
            
            # Extract items
            items = extractor.extract_items()
            
            # Generate keywords if OpenAI is enabled
            if self.use_openai:
                for item in items:
                    item.keywords = self.generate_keywords(item.description)
            
            self.all_items.extend(items)
            print(f"  Extracted {len(items)} items from {sheet_name}")
        
        self.workbook.close()
        return self.all_items
    
    def generate_keywords(self, description: str) -> List[str]:
        """Generate keywords using OpenAI API"""
        if not description or not OPENAI_API_KEY:
            return []
        
        try:
            client = openai.OpenAI(api_key=OPENAI_API_KEY)
            
            prompt = f"""Generate 5-8 relevant keywords for this construction item description.
            Keywords should be single words or short phrases that would help match this item in a bill of quantities.
            
            Description: {description}
            
            Return keywords as a comma-separated list."""
            
            response = client.chat.completions.create(
                model="gpt-3.5-turbo",
                messages=[
                    {"role": "system", "content": "You are a construction industry expert helping to categorize and tag construction items."},
                    {"role": "user", "content": prompt}
                ],
                max_tokens=100,
                temperature=0.3
            )
            
            keywords_text = response.choices[0].message.content.strip()
            keywords = [k.strip() for k in keywords_text.split(',')]
            return keywords[:8]  # Limit to 8 keywords
            
        except Exception as e:
            print(f"Error generating keywords: {e}")
            return []
    
    def export_to_csv(self, output_file: str = "pricelist_extracted.csv"):
        """Export extracted items to CSV"""
        if not self.all_items:
            print("No items to export")
            return
        
        # Prepare data for CSV
        csv_data = []
        for item in self.all_items:
            row = {
                'id': item.id,
                'code': item.code or '',
                'ref': item.ref or '',
                'description': item.description or '',
                'unit': item.unit or '',
                'category': item.category or '',
                'subCategory': item.subCategory or '',
                'keywords': '|'.join(item.keywords) if item.keywords else '',
            }
            
            # Add cell rates
            if item.cellRates:
                for i, (rate_key, rate_obj) in enumerate(item.cellRates.items(), 1):
                    row[f'cellRate{i}_reference'] = rate_obj.reference
                    row[f'cellRate{i}_sheetName'] = rate_obj.sheetName
                    row[f'cellRate{i}_rate'] = rate_obj.rate
            
            csv_data.append(row)
        
        # Write to CSV
        if csv_data:
            fieldnames = list(csv_data[0].keys())
            
            with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(csv_data)
            
            print(f"Exported {len(csv_data)} items to {output_file}")
    
    def export_to_json(self, output_file: str = "pricelist_extracted.json"):
        """Export extracted items to JSON"""
        if not self.all_items:
            print("No items to export")
            return
        
        # Convert items to dictionaries
        json_data = []
        for item in self.all_items:
            item_dict = {
                'id': item.id,
                'code': item.code,
                'ref': item.ref,
                'description': item.description,
                'unit': item.unit,
                'category': item.category,
                'subCategory': item.subCategory,
                'keywords': item.keywords,
            }
            
            # Convert cellRates
            if item.cellRates:
                cell_rates_dict = {}
                for rate_key, rate_obj in item.cellRates.items():
                    cell_rates_dict[rate_key] = {
                        'reference': rate_obj.reference,
                        'sheetName': rate_obj.sheetName,
                        'rate': rate_obj.rate
                    }
                item_dict['cellRates'] = cell_rates_dict
            else:
                item_dict['cellRates'] = None
            
            # Add patterns placeholder
            item_dict['patterns'] = item.patterns
            
            json_data.append(item_dict)
        
        # Write to JSON
        with open(output_file, 'w', encoding='utf-8') as jsonfile:
            json.dump(json_data, jsonfile, indent=2, ensure_ascii=False)
        
        print(f"Exported {len(json_data)} items to {output_file}")

def main():
    """Main execution function"""
    file_path = r"C:\Users\abaza\pricelist extraction\MJD-PRICELIST.xlsx"
    
    # Check if file exists
    if not Path(file_path).exists():
        print(f"Error: File not found at {file_path}")
        return
    
    # Initialize extractor
    print("="*50)
    print("MJD Pricelist Extraction Tool")
    print("="*50)
    
    # Ask user about OpenAI usage
    use_openai = False
    if OPENAI_API_KEY:
        response = input("OpenAI API key found. Generate keywords using AI? (y/n): ")
        use_openai = response.lower() == 'y'
    else:
        print("No OpenAI API key found. Set OPENAI_API_KEY environment variable to enable keyword generation.")
    
    # Create extractor
    extractor = PricelistExtractor(file_path, use_openai=use_openai)
    
    # Extract all sheets
    print("\nStarting extraction...")
    items = extractor.extract_all_sheets()
    
    print(f"\nTotal items extracted: {len(items)}")
    
    # Export results
    print("\nExporting results...")
    extractor.export_to_csv("pricelist_extracted.csv")
    extractor.export_to_json("pricelist_extracted.json")
    
    print("\nExtraction complete!")
    
    # Show sample of extracted data
    if items:
        print("\nSample of extracted items:")
        for item in items[:5]:
            print(f"  - {item.code}: {item.description[:50]}... [{item.category}]")

if __name__ == "__main__":
    main()