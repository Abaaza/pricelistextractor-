"""
Extraction script for External Works sheet
Handles the specific structure and format of the External Works pricelist
"""

from extractor_base import BaseExtractor
import pandas as pd
import json
import re
from openpyxl import load_workbook

class ExternalWorksExtractor(BaseExtractor):
    def __init__(self, excel_file='MJD-PRICELIST.xlsx'):
        super().__init__(excel_file, 'External Works')
        self.workbook = None
        self.worksheet = None
        self.current_subcategory = 'External Works'  # Default subcategory
        
    def load_workbook_for_formatting(self):
        """Load workbook with openpyxl to access formatting"""
        try:
            self.workbook = load_workbook(self.excel_file, data_only=True)
            self.worksheet = self.workbook[self.sheet_name]
            print(f"Loaded workbook for formatting detection")
        except Exception as e:
            print(f"Warning: Could not load workbook for formatting: {e}")
            self.worksheet = None
    
    def is_row_bold(self, row_idx):
        """Check if all non-empty cells in a row are bold"""
        if not self.worksheet:
            return False
        
        try:
            # Excel rows are 1-indexed
            excel_row = row_idx + 1
            row_is_bold = False
            has_content = False
            
            # Check first 5 columns for bold text
            for col in range(1, 6):
                cell = self.worksheet.cell(row=excel_row, column=col)
                if cell.value:
                    has_content = True
                    if cell.font and cell.font.bold:
                        row_is_bold = True
                    else:
                        # If any cell with content is not bold, row is not fully bold
                        return False
            
            return has_content and row_is_bold
        except Exception as e:
            return False
    
    def identify_data_rows(self):
        """Identify rows containing actual pricelist data"""
        data_rows = []
        
        for idx, row in self.df.iterrows():
            # Skip if row is mostly empty
            if row.notna().sum() < 3:
                continue
                
            # Look for patterns that indicate a data row
            # External Works typically has: Code | Description | Unit | Rate columns
            
            # Check if first column might be a code
            code_col = row[0] if 0 < len(row) else None
            if pd.notna(code_col):
                code_str = str(code_col).strip()
                # External Works codes often start with numbers or specific prefixes
                if (re.match(r'^\d+', code_str) or 
                    re.match(r'^[A-Z]\d+', code_str) or
                    re.match(r'^EW', code_str, re.I)):
                    data_rows.append(idx)
                    continue
            
            # Check if row has description-like content
            for col_idx in range(1, min(5, len(row))):
                cell = row[col_idx]
                if pd.notna(cell):
                    cell_str = str(cell).strip().lower()
                    # External Works keywords
                    if any(keyword in cell_str for keyword in 
                           ['paving', 'kerb', 'edging', 'fence', 'gate', 
                            'tarmac', 'asphalt', 'concrete', 'block', 'slab',
                            'drainage', 'channel', 'gulley', 'bollard', 'signage']):
                        data_rows.append(idx)
                        break
        
        return data_rows
    
    def extract_code(self, row, col_idx=0):
        """Extract code from row"""
        if col_idx < len(row) and pd.notna(row[col_idx]):
            code = str(row[col_idx]).strip()
            # Clean up code
            code = re.sub(r'\s+', '', code)  # Remove spaces
            if code and not code.lower() in ['nan', 'none', '-', '']:
                return code
        return None
    
    def extract_description(self, row, start_col=1):
        """Extract and clean description from columns B and C primarily"""
        description_parts = []
        
        # Column B (index 1) is usually the main description
        if len(row) > 1 and pd.notna(row[1]):
            desc = str(row[1]).strip()
            if desc and not self.is_unit(desc):
                description_parts.append(desc)
        
        # Column C (index 2) might have continuation or additional info
        if len(row) > 2 and pd.notna(row[2]):
            part = str(row[2]).strip()
            # Only add if it's not a unit and not a number
            if part and not self.is_unit(part) and not re.match(r'^[\d,\.]+$', part):
                description_parts.append(part)
        
        description = ' '.join(description_parts)
        
        # Clean common abbreviations
        replacements = {
            ' ne ': ' not exceeding ',
            ' n.e. ': ' not exceeding ',
            ' incl ': ' including ',
            ' excl ': ' excluding ',
            ' thk ': ' thick ',
            ' dp ': ' deep ',
            ' w ': ' wide ',
            ' h ': ' high ',
        }
        
        for old, new in replacements.items():
            description = description.replace(old, new)
        
        # Fix patterns
        description = re.sub(r'(\d+)thk', r'\1mm thick', description)
        description = re.sub(r'(\d+)dp', r'\1m deep', description)
        
        # Clean up spaces
        description = ' '.join(description.split())
        
        return description
    
    def is_unit(self, value):
        """Check if value is a unit"""
        if pd.isna(value):
            return False
        
        value_str = str(value).strip().lower()
        
        units = ['m', 'm2', 'm²', 'm3', 'm³', 'nr', 'no', 'item', 'sum',
                 'kg', 'tonnes', 't', 'lm', 'sqm', 'cum', 'each']
        
        return value_str in units
    
    def extract_unit(self, row):
        """Extract unit from row - primarily column E (index 4)"""
        # Check column E first (index 4) - this is the primary unit column
        if len(row) > 4 and pd.notna(row[4]):
            value = str(row[4]).strip()
            if self.is_unit(value):
                return self.standardize_unit(value)
        
        # Then check columns C and D as fallback
        for col_idx in [2, 3]:
            if col_idx < len(row) and pd.notna(row[col_idx]):
                value = str(row[col_idx]).strip()
                if self.is_unit(value):
                    return self.standardize_unit(value)
        
        # Infer from description if not found
        desc = self.extract_description(row)
        desc_lower = desc.lower()
        
        if any(word in desc_lower for word in ['paving', 'surfacing', 'tarmac']):
            return 'm2'
        elif any(word in desc_lower for word in ['kerb', 'edging', 'channel']):
            return 'm'
        elif any(word in desc_lower for word in ['bollard', 'sign', 'gate', 'post']):
            return 'nr'
        elif 'fence' in desc_lower or 'fencing' in desc_lower:
            return 'm'
        
        return 'item'
    
    def standardize_unit(self, unit):
        """Standardize unit format"""
        unit_map = {
            'm2': 'm2', 'sqm': 'm2', 'm²': 'm2',
            'm3': 'm3', 'cum': 'm3', 'm³': 'm3',
            'no': 'nr', 'no.': 'nr', 'each': 'nr',
            't': 'tonnes', 'tonne': 'tonnes',
            'lm': 'm', 'lin.m': 'm', 'l.m': 'm',
        }
        
        unit_lower = unit.lower()
        return unit_map.get(unit_lower, unit_lower)
    
    
    def extract_rate(self, row):
        """Extract rate value and column index"""
        # Look for rate in typical columns (starting from column F/index 5)
        for col_idx in range(5, min(15, len(row))):
            if pd.notna(row[col_idx]):
                value = str(row[col_idx]).strip()
                # Check if it's a number
                value_clean = value.replace(',', '').replace('£', '').replace('$', '')
                try:
                    rate = float(value_clean)
                    if rate > 0:  # Valid rate
                        return rate, col_idx
                except:
                    continue
        return None, None
    
    
    def determine_subcategory(self, description):
        """Determine subcategory based on description"""
        desc_lower = description.lower()
        
        # External Works subcategories
        if 'paving' in desc_lower or 'paved' in desc_lower:
            if 'block' in desc_lower:
                return 'Block Paving'
            elif 'slab' in desc_lower:
                return 'Slab Paving'
            else:
                return 'Paving'
        elif 'kerb' in desc_lower:
            return 'Kerbs and Edging'
        elif 'edging' in desc_lower:
            return 'Kerbs and Edging'
        elif 'fence' in desc_lower or 'fencing' in desc_lower:
            return 'Fencing'
        elif 'gate' in desc_lower:
            return 'Gates and Barriers'
        elif 'tarmac' in desc_lower or 'asphalt' in desc_lower:
            return 'Tarmac and Asphalt'
        elif 'concrete' in desc_lower:
            return 'Concrete Works'
        elif 'drainage' in desc_lower or 'gulley' in desc_lower or 'channel' in desc_lower:
            return 'Surface Drainage'
        elif 'bollard' in desc_lower:
            return 'Street Furniture'
        elif 'sign' in desc_lower or 'signage' in desc_lower:
            return 'Signage'
        elif 'marking' in desc_lower or 'line' in desc_lower:
            return 'Road Markings'
        else:
            return 'External Works'
    
    
    def generate_keywords(self, description):
        """Generate search keywords"""
        keywords = []
        desc_lower = description.lower()
        
        # Extract measurements
        measurements = re.findall(r'\d+(?:mm|m|kg|tonnes?)\b', desc_lower)
        keywords.extend(measurements[:2])
        
        # Key External Works terms
        terms = ['paving', 'kerb', 'edging', 'fence', 'gate', 
                 'tarmac', 'asphalt', 'concrete', 'bollard', 'drainage']
        
        for term in terms:
            if term in desc_lower:
                keywords.append(term)
        
        return keywords[:5]
    
    def create_item(self, row_idx, row, code, description, unit, subcategory, rate, rate_col_idx, keywords, current_id):
        """Create an item dictionary with proper formatting"""
        # Get cell references
        excel_ref = self.get_sheet_cell_reference(row_idx, 0)  # Reference to code cell
        rate_cell_ref = ''
        rate_value = 0.0
        
        if rate and rate_col_idx is not None:
            rate_cell_ref = self.get_sheet_cell_reference(row_idx, rate_col_idx)
            rate_value = rate
        
        # Create item matching Groundworks format
        item = {
            'id': current_id,  # Simple numeric ID
            'code': code if code else str(current_id),
            'description': description,
            'unit': unit,
            'category': 'External Works',
            'subcategory': subcategory,
            'rate': rate if rate else 0.0,
            'cellRate_reference': rate_cell_ref,
            'cellRate_rate': rate_value,
            'excelCellReference': excel_ref,
            'sourceSheetName': self.sheet_name,
            'keywords': ','.join(keywords) if keywords else ''
        }
        
        return item
    
    def extract_items(self):
        """Main extraction method with bold subcategory detection"""
        if self.df is None:
            self.load_sheet()
        
        # Load workbook for formatting detection
        self.load_workbook_for_formatting()
        
        print(f"\nExtracting items from {self.sheet_name}...")
        print(f"Starting from row 10...")
        
        items = []
        current_subcategory = 'External Works'  # Default subcategory
        rows_processed = 0
        rows_skipped = 0
        
        # Process all rows starting from row 10
        start_row = 9  # Row 10 in Excel (0-indexed)
        
        for row_idx in range(start_row, len(self.df)):
            row = self.df.iloc[row_idx]
            
            # Skip if row is completely empty
            if row.notna().sum() < 1:
                continue
            
            # Check if this row is bold (potential subcategory header)
            if self.is_row_bold(row_idx):
                # Extract text from the row to use as subcategory
                subcategory_text = ''
                for col_idx in range(min(5, len(row))):
                    if pd.notna(row[col_idx]):
                        text = str(row[col_idx]).strip()
                        if text and not self.is_unit(text):
                            subcategory_text = text
                            break
                
                if subcategory_text:
                    current_subcategory = subcategory_text
                    print(f"Found subcategory at row {row_idx + 1}: {current_subcategory}")
                    continue  # Skip this row as it's a header
            
            # Check if this is a data row
            first_col = row[0] if 0 < len(row) else None
            second_col = row[1] if len(row) > 1 else None
            
            # Skip if column A is empty
            if pd.isna(first_col):
                continue
                
            # Skip if column B is empty or too short
            if pd.isna(second_col):
                continue
                
            first_str = str(first_col).strip()
            second_str = str(second_col).strip()
            
            # Skip if column A has nothing meaningful
            if not first_str or first_str.lower() in ['', 'nan', 'none']:
                continue
                
            # Skip if column B is too short to be a description
            if len(second_str) < 5:
                continue
                
            # Skip if column B is just a unit
            if self.is_unit(second_str):
                continue
            
            rows_processed += 1
            
            # Extract code (actual Excel value)
            code = self.extract_code(row)
            
            # Extract description
            description = self.extract_description(row)
            
            # Skip if no valid description
            if not description or len(description) < 5:
                rows_skipped += 1
                continue
            
            # Extract unit
            unit = self.extract_unit(row)
            
            # Extract rate and column index
            rate, rate_col_idx = self.extract_rate(row)
            
            # Use current subcategory (from bold header) or determine from keywords as fallback
            if current_subcategory and current_subcategory != 'External Works':
                subcategory = current_subcategory
            else:
                subcategory = self.determine_subcategory(description)
            
            # Generate keywords
            keywords = self.generate_keywords(description)
            
            # Create item with actual code
            item = self.create_item(
                row_idx=row_idx,
                row=row,
                code=code,
                description=description,
                unit=unit,
                subcategory=subcategory,
                rate=rate,
                rate_col_idx=rate_col_idx,
                keywords=keywords,
                current_id=len(items) + 1
            )
            
            items.append(item)
        
        self.extracted_items = items
        print(f"Processed {rows_processed} rows, skipped {rows_skipped} due to short descriptions")
        print(f"Extracted {len(items)} valid items from {self.sheet_name}")
        return items
    
    def save_output(self, output_prefix='external_works'):
        """Save extracted data"""
        if not self.extracted_items:
            print("No items to save")
            return
        
        # Save JSON
        json_file = f"{output_prefix}_extracted.json"
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(self.extracted_items, f, indent=2, ensure_ascii=False)
        print(f"Saved JSON: {json_file}")
        
        # Save CSV
        df = pd.DataFrame(self.extracted_items)
        # Keywords are already formatted as comma-separated strings
        csv_file = f"{output_prefix}_extracted.csv"
        df.to_csv(csv_file, index=False)
        print(f"Saved CSV: {csv_file}")
        
        return json_file, csv_file

def main():
    print("="*60)
    print("EXTERNAL WORKS SHEET EXTRACTION")
    print("="*60)
    
    extractor = ExternalWorksExtractor()
    items = extractor.extract_items()
    
    if items:
        # Show sample
        print("\nSample extracted items:")
        for item in items[:3]:
            print(f"\nID: {item['id']}")
            print(f"  Code: {item['code']}")
            print(f"  Description: {item['description'][:60]}...")
            print(f"  Unit: {item['unit']}")
            print(f"  Subcategory: {item['subcategory']}")
            print(f"  Rate: {item['rate']}")
            print(f"  Cell Ref: {item['cellRate_reference']}")
            print(f"  Keywords: {', '.join(item['keywords'][:3])}")
        
        extractor.save_output()
        
        # Statistics
        print("\n" + "="*60)
        print("EXTRACTION STATISTICS")
        print("="*60)
        print(f"Total items: {len(items)}")
        print(f"Items with rates: {sum(1 for i in items if i['rate'])}")
        print(f"Items with cell references: {sum(1 for i in items if i['cellRate_reference'])}")
        
        # Subcategory distribution
        subcats = {}
        for item in items:
            subcat = item['subcategory']
            subcats[subcat] = subcats.get(subcat, 0) + 1
        
        print("\nSubcategory distribution:")
        for subcat, count in sorted(subcats.items(), key=lambda x: x[1], reverse=True)[:5]:
            print(f"  {subcat}: {count}")

if __name__ == "__main__":
    main()