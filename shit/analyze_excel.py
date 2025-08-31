import pandas as pd
import openpyxl
from pathlib import Path

# Load the Excel file
file_path = r"C:\Users\abaza\pricelist extraction\MJD-PRICELIST.xlsx"
workbook = openpyxl.load_workbook(file_path, data_only=True)

print(f"Total sheets: {len(workbook.sheetnames)}")
print(f"Sheet names: {workbook.sheetnames}")
print("\n" + "="*50 + "\n")

# Analyze each sheet structure
for sheet_name in workbook.sheetnames[:5]:  # First 5 sheets for initial analysis
    print(f"Sheet: {sheet_name}")
    print("-" * 30)
    
    sheet = workbook[sheet_name]
    
    # Get dimensions
    max_row = sheet.max_row
    max_col = sheet.max_column
    print(f"Dimensions: {max_row} rows × {max_col} columns")
    
    # Sample first 10 rows to understand structure
    print("First 10 rows sample:")
    for row in sheet.iter_rows(min_row=1, max_row=min(10, max_row), values_only=True):
        # Filter out completely empty rows
        if any(cell is not None for cell in row):
            print(row[:10])  # Show first 10 columns
    
    print("\n")

workbook.close()