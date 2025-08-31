"""
Specialized Drainage Sheet Extractor
Extracts drainage items with combined header + range descriptions
Following the schema from pricelist_complete_finallll.csv
"""

import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd
import json
from typing import List, Dict, Optional, Tuple
from dataclasses import dataclass, asdict
import re

@dataclass
class DrainageItem:
    """Schema matching the complete pricelist structure"""
    id: str
    code: str
    original_code: str
    description: str
    unit: str
    category: str
    subcategory: str
    work_type: str
    rate: Optional[float] = 0.0
    cellRate_reference: Optional[str] = None
    cellRate_rate: Optional[float] = 0.0
    excelCellReference: Optional[str] = None
    sourceSheetName: str = "Drainage"
    keywords: Optional[str] = None

class DrainageExtractor:
    """Extract drainage items with range-based descriptions"""
    
    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.workbook = None
        self.worksheet = None
        self.items = []
        
    def load_sheet(self):
        """Load the Drainage worksheet"""
        self.workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
        self.worksheet = self.workbook['Drainage']
        print(f"Loaded Drainage sheet with {self.worksheet.max_row} rows")
        
    def is_header_row(self, row: int) -> Optional[str]:
        """Check if a row contains a header description"""
        desc = self.worksheet.cell(row=row, column=2).value
        if desc and isinstance(desc, str) and len(desc) > 20:
            # Common patterns in drainage headers
            keywords = ['excavat', 'trench', 'pipe', 'backfill', 'dispose', 
                       'manhole', 'gulley', 'drain', 'sewer']
            if any(keyword in desc.lower() for keyword in keywords):
                return desc
        return None
    
    def get_range_description(self, row: int) -> Optional[Tuple[str, str, str]]:
        """Extract range information from a row"""
        col_c = self.worksheet.cell(row=row, column=3).value  # Start of range or 'ne'
        col_d = self.worksheet.cell(row=row, column=4).value  # Usually '-'
        col_e = self.worksheet.cell(row=row, column=5).value  # End of range
        col_f = self.worksheet.cell(row=row, column=6).value  # Unit
        
        # Handle 'ne' (not exceeding) cases
        if col_c == 'ne' and col_e is not None:
            range_desc = f"not exceeding {col_e}"
        elif col_c is not None and col_e is not None:
            range_desc = f"{col_c} - {col_e}"
        else:
            return None
            
        # Get unit
        unit = col_f if col_f else 'm'
        
        return range_desc, unit, str(col_d) if col_d else '-'
    
    def determine_subcategory(self, description: str) -> str:
        """Determine subcategory based on description"""
        desc_lower = description.lower()
        
        if 'manhole' in desc_lower:
            return 'Manholes'
        elif 'gulley' in desc_lower or 'gully' in desc_lower:
            return 'Gullies'
        elif 'below ground' in desc_lower:
            return 'Below Ground Drainage'
        elif 'above ground' in desc_lower:
            return 'Above Ground Drainage'
        elif 'pipe' in desc_lower or 'pipework' in desc_lower:
            return 'Pipework'
        elif 'excavat' in desc_lower:
            return 'Excavation'
        elif 'backfill' in desc_lower:
            return 'Backfilling'
        else:
            return 'General Drainage'
    
    def extract_keywords(self, description: str) -> str:
        """Extract keywords from description"""
        keywords = []
        
        # Extract pipe sizes
        pipe_sizes = re.findall(r'\d+mm', description.lower())
        keywords.extend(pipe_sizes)
        
        # Common drainage keywords
        keyword_patterns = ['excavate', 'trench', 'pipe', 'backfill', 'sem', 
                          'dispose', 'manhole', 'gulley', 'drainage', 'sewer',
                          'invert', 'depth']
        
        for pattern in keyword_patterns:
            if pattern in description.lower():
                keywords.append(pattern)
        
        return '|'.join(keywords[:8])  # Limit to 8 keywords
    
    def extract_items(self):
        """Extract all drainage items with combined descriptions"""
        print("\nExtracting drainage items...")
        
        current_header = None
        item_counter = 1
        
        for row in range(1, self.worksheet.max_row + 1):
            # Check if this row is a header
            header = self.is_header_row(row)
            if header:
                current_header = header
                print(f"\nFound header at row {row}: {header[:60]}...")
                continue
            
            # Skip if no current header
            if not current_header:
                continue
            
            # Check if this row has range data
            item_code = self.worksheet.cell(row=row, column=1).value
            if not item_code or not isinstance(item_code, (int, float)):
                continue
                
            range_info = self.get_range_description(row)
            if not range_info:
                continue
            
            range_desc, unit, separator = range_info
            
            # Combine header with range to create full description
            full_description = f"{current_header}; depth to invert: {range_desc} {unit}"
            
            # Clean up the description
            full_description = re.sub(r'\s+', ' ', full_description)  # Remove extra spaces
            full_description = full_description.replace(';;', ';')    # Remove double semicolons
            
            # Generate item ID and code
            item_id = f"DR_{item_counter:04d}_Drainage"
            code = f"DR{item_counter:04d}"
            
            # Get rate from column O (column 15)
            rate_value = self.worksheet.cell(row=row, column=15).value
            rate = 0.0
            rate_cell_ref = None
            
            if rate_value and isinstance(rate_value, (int, float)) and rate_value > 0:
                rate = float(rate_value)
                rate_cell_ref = f"Drainage!{get_column_letter(15)}{row}"
            else:
                # Try column T (20) as alternative (NET ALLOWANCE)
                rate_value = self.worksheet.cell(row=row, column=20).value
                if rate_value and isinstance(rate_value, (int, float)) and rate_value > 0:
                    rate = float(rate_value)
                    rate_cell_ref = f"Drainage!{get_column_letter(20)}{row}"
            
            # Get cell reference for the item
            cell_ref = f"Drainage!{get_column_letter(1)}{row}"
            
            # Determine subcategory and work type
            subcategory = self.determine_subcategory(full_description)
            work_type = "Drainage Works"
            
            # Extract keywords
            keywords = self.extract_keywords(full_description)
            
            # Create item
            item = DrainageItem(
                id=item_id,
                code=code,
                original_code=str(item_code),
                description=full_description,
                unit=unit,
                category="Drainage",
                subcategory=subcategory,
                work_type=work_type,
                rate=rate,
                cellRate_reference=rate_cell_ref,
                cellRate_rate=rate,
                excelCellReference=cell_ref,
                sourceSheetName="Drainage",
                keywords=keywords
            )
            
            self.items.append(item)
            item_counter += 1
        
        print(f"\nExtracted {len(self.items)} drainage items")
        return self.items
    
    def save_to_csv(self, filename: str = "drainage.csv"):
        """Save extracted items to CSV"""
        if not self.items:
            print("No items to save")
            return
        
        # Convert to dataframe
        df_data = []
        for item in self.items:
            row = asdict(item)
            df_data.append(row)
        
        df = pd.DataFrame(df_data)
        
        # Save to CSV
        df.to_csv(filename, index=False)
        print(f"Saved {len(self.items)} items to {filename}")
        
        # Show sample with rates
        print("\nSample items with rates:")
        for item in self.items[:5]:
            desc = item.description[:70] + "..." if len(item.description) > 70 else item.description
            print(f"  {item.code}: {desc}")
            print(f"    Rate: £{item.rate:.2f} (Cell: {item.cellRate_reference if item.cellRate_reference else 'No rate'})")
    
    def save_to_json(self, filename: str = "drainage.json"):
        """Save extracted items to JSON"""
        if not self.items:
            print("No items to save")
            return
        
        # Convert to list of dictionaries
        json_data = []
        for item in self.items:
            json_data.append(asdict(item))
        
        # Save to JSON
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, indent=2, ensure_ascii=False)
        
        print(f"Saved {len(self.items)} items to {filename}")
    
    def close(self):
        """Close the workbook"""
        if self.workbook:
            self.workbook.close()

def main():
    """Main execution"""
    excel_path = "MJD-PRICELIST.xlsx"
    
    print("="*60)
    print("DRAINAGE SHEET EXTRACTOR")
    print("="*60)
    
    extractor = DrainageExtractor(excel_path)
    
    try:
        # Load and extract
        extractor.load_sheet()
        items = extractor.extract_items()
        
        # Save results
        print("\nSaving results...")
        extractor.save_to_csv("drainage.csv")
        extractor.save_to_json("drainage.json")
        
        # Show statistics
        print("\n" + "="*60)
        print("EXTRACTION COMPLETE")
        print("="*60)
        print(f"Total items extracted: {len(items)}")
        
        # Count by subcategory
        subcategories = {}
        for item in items:
            subcat = item.subcategory
            subcategories[subcat] = subcategories.get(subcat, 0) + 1
        
        print("\nItems by subcategory:")
        for subcat, count in sorted(subcategories.items()):
            print(f"  {subcat}: {count} items")
        
        # Count items with rates
        items_with_rates = sum(1 for item in items if item.rate > 0)
        items_without_rates = len(items) - items_with_rates
        
        print(f"\nRate statistics:")
        print(f"  Items with rates: {items_with_rates} ({items_with_rates/len(items)*100:.1f}%)")
        print(f"  Items without rates: {items_without_rates} ({items_without_rates/len(items)*100:.1f}%)")
        
        if items_with_rates > 0:
            rates = [item.rate for item in items if item.rate > 0]
            print(f"  Rate range: £{min(rates):.2f} - £{max(rates):.2f}")
            print(f"  Average rate: £{sum(rates)/len(rates):.2f}")
        
    finally:
        extractor.close()

if __name__ == "__main__":
    main()