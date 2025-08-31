"""
MJD Pricelist Extraction V2 - Preserves Original Codes with Qwen Enhancement
Maintains exact codes from Excel while enhancing quality
"""

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import numpy as np
import json
import csv
import re
from typing import Dict, List, Optional, Tuple, Any
from dataclasses import dataclass, asdict, field
from pathlib import Path
import os
from datetime import datetime
import time
import hashlib

# DeepInfra Configuration
DEEPINFRA_API_KEY = "8MSsOohjJBtIAlzstuh4inhRzgnuS68k"
DEEPINFRA_BASE_URL = "https://api.deepinfra.com/v1/openai"
DEEPINFRA_MODEL = "Qwen/Qwen2.5-72B-Instruct"

try:
    import openai
    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False
    print("Warning: OpenAI library not installed. Run: pip install openai")

@dataclass
class CellRate:
    """Simplified cell rate reference"""
    reference: str  # e.g. "Groundworks!C14"
    rate: float

@dataclass
class PriceItem:
    """Schema matching the database structure exactly"""
    # Core identification fields
    id: str                                    # Unique identifier
    description: str                           # Item description (required)
    
    # Optional fields
    code: Optional[str] = None                # Item code FROM EXCEL
    ref: Optional[str] = None                 # Reference number
    original_code: Optional[str] = None       # Original code from sheet
    keywords: Optional[List[str]] = None      # Keywords for searching
    patterns: Optional[List[Dict]] = None     # Learning patterns
    
    # Cell-based rate reference (SIMPLIFIED)
    cellRate: Optional[Dict] = None           # Single cell reference
    
    # Essential categorization fields
    category: Optional[str] = None            # Main category
    subcategory: Optional[str] = None         # Subcategory  
    work_type: Optional[str] = None           # Type of work
    brand: Optional[str] = None               # Brand if applicable
    unit: Optional[str] = None                # Unit of measurement
    rate: Optional[float] = None              # Primary rate
    
    # Excel mapping fields for rate updates
    excelCellReference: Optional[str] = None  # e.g., "Groundworks!F11"
    sourceSheetName: Optional[str] = None     # e.g., "Groundworks"
    sourceRowNumber: Optional[int] = None     # e.g., 11
    sourceColumnLetter: Optional[str] = None  # e.g., "F"
    
    # Legacy fields (keeping for compatibility)
    subCategoryCode: Optional[str] = None
    subCategoryName: Optional[str] = None
    sub_category: Optional[str] = None        # Alternative subcategory field
    
    # Metadata
    isActive: bool = True                     # Whether item is active
    createdAt: int = 0                        # Creation timestamp
    updatedAt: int = 0                        # Last update timestamp
    createdBy: str = "system"                 # User who created the item

class QwenEnhancer:
    """Handles DeepInfra Qwen API calls for quality improvement"""
    
    def __init__(self, api_key: str = DEEPINFRA_API_KEY):
        self.api_key = api_key
        if api_key and OPENAI_AVAILABLE:
            self.client = openai.OpenAI(
                api_key=api_key,
                base_url=DEEPINFRA_BASE_URL
            )
            print("DeepInfra Qwen 72B initialized successfully")
        else:
            self.client = None
        self.api_calls = 0
        self.cache = {}
    
    def enhance_batch(self, items: List[Dict], sheet_name: str) -> List[Dict]:
        """Enhance items with Qwen while preserving original codes"""
        if not self.client:
            return items
        
        batch_size = 8  # Optimal for Qwen
        enhanced_items = []
        total_batches = (len(items) + batch_size - 1) // batch_size
        
        for i in range(0, len(items), batch_size):
            batch_num = (i // batch_size) + 1
            print(f"    Enhancing batch {batch_num}/{total_batches}...")
            
            batch = items[i:i + batch_size]
            try:
                enhanced_batch = self._process_batch(batch, sheet_name)
                enhanced_items.extend(enhanced_batch)
            except Exception as e:
                print(f"      Batch failed: {str(e)[:50]}, using original")
                enhanced_items.extend(batch)
            
            if batch_num < total_batches:
                time.sleep(0.3)  # Rate limiting
        
        return enhanced_items
    
    def _process_batch(self, batch: List[Dict], sheet_name: str) -> List[Dict]:
        """Process batch with Qwen"""
        # Prepare data
        batch_data = []
        for idx, item in enumerate(batch):
            batch_data.append({
                'index': idx,
                'code': item.get('original_code', ''),
                'description': item.get('description', ''),
                'unit': item.get('unit', ''),
                'rate': item.get('rate', '')
            })
        
        prompt = f"""You are reviewing construction pricelist items from "{sheet_name}" sheet.

Items to enhance:
{json.dumps(batch_data, indent=2)}

For each item, improve quality while PRESERVING original data:

1. **Description**: 
   - Fix truncation and typos
   - Complete abbreviated text (e.g., "ne" = "not exceeding", "thk" = "thick")
   - Keep measurements and specifications exact
   - Preserve technical codes (BS, EN, etc.)

2. **Unit**: Correct to standard construction units:
   m, m², m³, nr, kg, tonnes, hour, day, week, month, sum, item, %

3. **Work Type** based on description:
   Excavation, Concrete, Formwork, Reinforcement, Masonry, Drainage, 
   Electrical, Mechanical, Plumbing, HVAC, Finishes, Painting, Roofing,
   External Works, Landscaping, Preliminaries, Temporary Works

4. **Subcategory** for {sheet_name}:
   - Groundworks: Site Clearance, Excavation, Earthworks Support, Disposal, Filling
   - RC Works: In-situ Concrete, Formwork, Reinforcement, Precast Concrete
   - Drainage: Below Ground Drainage, Manholes, Pipework, Gullies
   - External Works: Paving, Kerbs, Fencing, Landscaping
   - Services: Electrical, Mechanical, Plumbing, HVAC

5. **Keywords**: 4-6 specific search terms

IMPORTANT: Keep original code unchanged!

Return JSON array:
[{{
  "index": 0,
  "code": "keep original",
  "description": "enhanced description",
  "unit": "standard unit",
  "work_type": "type",
  "subcategory": "subcategory",
  "keywords": ["keyword1", "keyword2", "keyword3"]
}}]"""

        try:
            response = self.client.chat.completions.create(
                model=DEEPINFRA_MODEL,
                messages=[
                    {"role": "system", "content": "You are a construction estimator. Return only valid JSON."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.2,
                max_tokens=2000
            )
            
            self.api_calls += 1
            result_text = response.choices[0].message.content.strip()
            
            # Parse JSON
            if '```json' in result_text:
                result_text = result_text.split('```json')[1].split('```')[0]
            elif '```' in result_text:
                result_text = result_text.split('```')[1].split('```')[0]
            
            enhanced_data = json.loads(result_text)
            
            # Merge enhancements
            enhanced_batch = []
            for item in batch:
                enhanced_item = item.copy()
                
                # Find matching enhancement
                for enhanced in enhanced_data:
                    if enhanced.get('index') == batch.index(item):
                        # Apply enhancements but keep original code
                        enhanced_item['description'] = enhanced.get('description', item['description'])
                        enhanced_item['unit'] = enhanced.get('unit', item.get('unit'))
                        enhanced_item['work_type'] = enhanced.get('work_type')
                        enhanced_item['subcategory'] = enhanced.get('subcategory')
                        enhanced_item['keywords'] = enhanced.get('keywords', [])
                        break
                
                enhanced_batch.append(enhanced_item)
            
            return enhanced_batch
            
        except Exception as e:
            print(f"      API error: {str(e)[:50]}")
            return batch

class ImprovedSheetExtractor:
    """Base extractor with better code preservation"""
    
    def __init__(self, sheet_name: str, worksheet, enhancer: Optional[QwenEnhancer] = None):
        self.sheet_name = sheet_name
        self.worksheet = worksheet
        self.enhancer = enhancer
        self.items = []
        self.current_timestamp = int(datetime.now().timestamp() * 1000)
        self.seen_items = set()  # Track unique items
    
    def get_full_cell_reference(self, row: int, col: int) -> str:
        """Get full cell reference with sheet name"""
        return f"{self.sheet_name}!{get_column_letter(col)}{row}"
    
    def clean_value(self, value: Any) -> Any:
        """Clean cell values"""
        if value is None:
            return None
        if isinstance(value, str):
            cleaned = ' '.join(value.split())
            return cleaned
        return value
    
    def extract_code(self, row: int) -> Optional[str]:
        """Extract code from column A or other likely locations"""
        # Check column A first (most common for codes)
        code = self.worksheet.cell(row=row, column=1).value
        
        if code is not None:
            code_str = str(code).strip()
            
            # Check if it's a valid code (not a header or description)
            if code_str and len(code_str) < 20:
                # Could be numeric (1, 2, 3) or alphanumeric (A1, RC01, etc.)
                if code_str.isdigit() or any(c.isdigit() for c in code_str):
                    return code_str
                # Could be a reference like "Item 1" or "Ref A"
                if len(code_str) < 10:
                    return code_str
        
        return None
    
    def is_valid_rate(self, value: Any) -> bool:
        """Check if value is a valid rate"""
        if value is None:
            return False
        if isinstance(value, str):
            skip_values = ['by others', 'included', 'n/a', 'nil', '-', 'tbc', 'item']
            if any(skip in value.lower() for skip in skip_values):
                return False
        try:
            rate = float(value)
            return 0 < rate < 1000000
        except:
            return False
    
    def is_valid_description(self, desc: Any) -> bool:
        """Check if description is valid"""
        if not desc:
            return False
        
        desc = str(desc) if not isinstance(desc, str) else desc
        
        if len(desc) < 3:
            return False
        
        if not any(c.isalpha() for c in desc):
            return False
        
        # Skip headers and totals
        skip_terms = ['total', 'subtotal', 'carried forward', 'brought forward',
                     'continued', 'blank', 'n/a', 'nil', 'schedule of']
        
        desc_lower = desc.lower()
        return not any(term in desc_lower for term in skip_terms)
    
    def create_unique_id(self, code: str, sheet_name: str, row: int) -> str:
        """Create unique ID preserving original code"""
        # Use sheet prefix + original code
        sheet_prefix = ''.join([c for c in sheet_name.replace(' ', '').upper() if c.isalpha()])[:3]
        
        if code:
            # If we have original code, use it
            base_id = f"{sheet_prefix}_{code}"
        else:
            # Fallback to row-based ID
            base_id = f"{sheet_prefix}_R{row}"
        
        # Ensure uniqueness
        if base_id in self.seen_items:
            counter = 1
            while f"{base_id}_{counter}" in self.seen_items:
                counter += 1
            base_id = f"{base_id}_{counter}"
        
        self.seen_items.add(base_id)
        return base_id
    
    def extract_items(self) -> List[PriceItem]:
        """Extract items preserving original codes"""
        items = []
        current_subcategory = None
        
        # Find data start row
        data_start_row = 10
        for row in range(1, min(30, self.worksheet.max_row + 1)):
            cell = self.worksheet.cell(row=row, column=2)
            if cell.value and 'description' in str(cell.value).lower():
                data_start_row = row + 2
                break
        
        # Track codes to avoid duplicates
        used_codes = set()
        
        # Process rows
        for row in range(data_start_row, self.worksheet.max_row + 1):
            # Check for subcategory headers
            cell_b = self.worksheet.cell(row=row, column=2)
            
            if cell_b.value:
                cell_value = self.clean_value(cell_b.value)
                
                # Check if it's a subcategory header
                if cell_b.font and cell_b.font.bold and len(str(cell_value)) < 50:
                    current_subcategory = cell_value
                    continue
                
                # Common subcategory patterns
                if cell_value and isinstance(cell_value, str) and cell_value.isupper() and len(cell_value) < 50:
                    # Check if it's not an item description
                    if not any(char.isdigit() for char in cell_value):
                        current_subcategory = cell_value
                        continue
            
            # Extract item data
            description = self.clean_value(cell_b.value)
            
            if not self.is_valid_description(description):
                continue
            
            # Extract original code
            original_code = self.extract_code(row)
            
            # Get other data
            quantity = self.worksheet.cell(row=row, column=4).value
            unit = self.clean_value(self.worksheet.cell(row=row, column=5).value)
            rate = self.worksheet.cell(row=row, column=6).value
            
            # Find valid rate column
            rate_col = 6
            if not self.is_valid_rate(rate):
                for col in [7, 8, 9, 10]:
                    alt_rate = self.worksheet.cell(row=row, column=col).value
                    if self.is_valid_rate(alt_rate):
                        rate = alt_rate
                        rate_col = col
                        break
            
            # Process valid items
            if description and (rate or quantity):
                # Create unique ID preserving original code
                item_id = self.create_unique_id(original_code, self.sheet_name, row)
                
                # Use original code or create one
                if original_code:
                    item_code = original_code
                else:
                    # Generate sequential code if not present
                    item_code = str(len(items) + 1)
                
                # Ensure code uniqueness within sheet
                if item_code in used_codes:
                    base_code = item_code
                    counter = 1
                    while f"{base_code}_{counter}" in used_codes:
                        counter += 1
                    item_code = f"{base_code}_{counter}"
                
                used_codes.add(item_code)
                
                # Create cell rate reference
                cell_rate = None
                if self.is_valid_rate(rate):
                    cell_rate = {
                        'reference': self.get_full_cell_reference(row, rate_col),
                        'rate': float(rate)
                    }
                
                # Auto-detect subcategory if not set
                if not current_subcategory:
                    desc_lower = description.lower()
                    # Sheet-specific subcategory detection
                    if 'concrete' in desc_lower:
                        current_subcategory = 'Concrete Works'
                    elif 'excavat' in desc_lower:
                        current_subcategory = 'Excavation'
                    elif 'formwork' in desc_lower:
                        current_subcategory = 'Formwork'
                    elif 'reinforc' in desc_lower or 'rebar' in desc_lower:
                        current_subcategory = 'Reinforcement'
                    elif 'drain' in desc_lower:
                        current_subcategory = 'Drainage'
                    else:
                        current_subcategory = f"General {self.sheet_name}"
                
                # Create price item
                item = PriceItem(
                    id=item_id,
                    code=item_code,
                    original_code=original_code,
                    description=description,
                    unit=unit,
                    rate=float(rate) if self.is_valid_rate(rate) else None,
                    category=self.sheet_name,
                    subcategory=current_subcategory,
                    sub_category=current_subcategory,
                    cellRate=cell_rate,
                    excelCellReference=self.get_full_cell_reference(row, rate_col) if rate else None,
                    sourceSheetName=self.sheet_name,
                    sourceRowNumber=row,
                    sourceColumnLetter=get_column_letter(rate_col) if rate else None,
                    isActive=True,
                    createdAt=self.current_timestamp,
                    updatedAt=self.current_timestamp,
                    createdBy="system"
                )
                
                items.append(item)
        
        # Enhance with Qwen if available
        if self.enhancer and items:
            items = self.post_process_items(items)
        
        return items
    
    def post_process_items(self, items: List[PriceItem]) -> List[PriceItem]:
        """Enhance items with Qwen"""
        print(f"  Enhancing {len(items)} items with Qwen 72B...")
        
        # Convert to dicts
        items_dict = []
        for item in items:
            items_dict.append({
                'description': item.description,
                'unit': item.unit,
                'subcategory': item.subcategory,
                'rate': item.rate,
                'original_code': item.original_code,
                'code': item.code
            })
        
        # Enhance
        enhanced_dicts = self.enhancer.enhance_batch(items_dict, self.sheet_name)
        
        # Apply enhancements
        for i, item in enumerate(items):
            if i < len(enhanced_dicts):
                enhanced = enhanced_dicts[i]
                
                # Update fields while preserving codes
                item.description = enhanced.get('description', item.description)
                item.unit = enhanced.get('unit', item.unit)
                item.work_type = enhanced.get('work_type')
                item.subcategory = enhanced.get('subcategory', item.subcategory)
                item.sub_category = item.subcategory
                item.keywords = enhanced.get('keywords', [])
        
        return items

class ImprovedPricelistExtractor:
    """Main extractor with deduplication and original code preservation"""
    
    SKIP_SHEETS = ['Summary', 'Set factors & prices', 'Tender Summary', 'Budget Costings']
    
    def __init__(self, file_path: str, use_qwen: bool = True):
        self.file_path = file_path
        self.use_qwen = use_qwen
        self.workbook = None
        self.all_items = []
        self.unique_items = {}  # For deduplication
        
        if use_qwen:
            self.enhancer = QwenEnhancer(DEEPINFRA_API_KEY)
        else:
            self.enhancer = None
    
    def extract_all_sheets(self) -> List[PriceItem]:
        """Extract from all sheets with deduplication"""
        self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
        
        print(f"\nProcessing {len(self.workbook.sheetnames)} sheets")
        print("="*60)
        
        for sheet_name in self.workbook.sheetnames:
            if sheet_name in self.SKIP_SHEETS:
                print(f"Skipping: {sheet_name}")
                continue
            
            print(f"\nProcessing: {sheet_name}")
            worksheet = self.workbook[sheet_name]
            
            # Use improved extractor
            extractor = ImprovedSheetExtractor(sheet_name, worksheet, self.enhancer)
            items = extractor.extract_items()
            
            # Deduplicate
            for item in items:
                # Create unique key
                unique_key = f"{item.description}|{item.category}|{item.unit}"
                
                if unique_key not in self.unique_items:
                    self.unique_items[unique_key] = item
                    self.all_items.append(item)
                else:
                    # Keep the one with original code
                    existing = self.unique_items[unique_key]
                    if item.original_code and not existing.original_code:
                        self.unique_items[unique_key] = item
                        # Replace in list
                        idx = self.all_items.index(existing)
                        self.all_items[idx] = item
            
            print(f"  Extracted: {len(items)} items (unique total: {len(self.all_items)})")
        
        self.workbook.close()
        
        # Print summary
        self.print_summary()
        
        return self.all_items
    
    def print_summary(self):
        """Print extraction summary"""
        print("\n" + "="*60)
        print("EXTRACTION SUMMARY")
        print("="*60)
        print(f"Total unique items: {len(self.all_items)}")
        
        # Count by category
        categories = {}
        items_with_codes = 0
        items_with_keywords = 0
        
        for item in self.all_items:
            cat = item.category or "Unknown"
            categories[cat] = categories.get(cat, 0) + 1
            
            if item.original_code:
                items_with_codes += 1
            if item.keywords:
                items_with_keywords += 1
        
        print("\nItems by category:")
        for cat, count in sorted(categories.items(), key=lambda x: x[1], reverse=True)[:10]:
            print(f"  {cat}: {count}")
        
        print(f"\nQuality metrics:")
        print(f"  Items with original codes: {items_with_codes}/{len(self.all_items)}")
        print(f"  Items with keywords: {items_with_keywords}/{len(self.all_items)}")
        
        if self.enhancer:
            print(f"  API calls made: {self.enhancer.api_calls}")
    
    def export_to_csv(self, output_file: str = "pricelist_v2.csv"):
        """Export to CSV"""
        if not self.all_items:
            print("No items to export")
            return
        
        csv_data = []
        for item in self.all_items:
            row = {
                'id': item.id,
                'code': item.code,
                'original_code': item.original_code or '',
                'description': item.description,
                'keywords': '|'.join(item.keywords) if item.keywords else '',
                'category': item.category or '',
                'subcategory': item.subcategory or '',
                'work_type': item.work_type or '',
                'unit': item.unit or '',
                'rate': item.rate or '',
                'cellRate_reference': item.cellRate['reference'] if item.cellRate else '',
                'cellRate_rate': item.cellRate['rate'] if item.cellRate else '',
                'excelCellReference': item.excelCellReference or '',
                'sourceSheetName': item.sourceSheetName or '',
                'sourceRowNumber': item.sourceRowNumber or '',
                'sourceColumnLetter': item.sourceColumnLetter or '',
                'isActive': item.isActive,
                'createdAt': item.createdAt,
                'updatedAt': item.updatedAt
            }
            csv_data.append(row)
        
        # Write CSV
        fieldnames = list(csv_data[0].keys()) if csv_data else []
        
        with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(csv_data)
        
        print(f"\nExported {len(csv_data)} items to {output_file}")
    
    def export_to_json(self, output_file: str = "pricelist_v2.json"):
        """Export to JSON"""
        if not self.all_items:
            return
        
        json_data = []
        for item in self.all_items:
            item_dict = asdict(item)
            json_data.append(item_dict)
        
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, indent=2, ensure_ascii=False)
        
        print(f"Exported {len(json_data)} items to {output_file}")

def main():
    """Main execution"""
    file_path = r"C:\Users\abaza\pricelist extraction\MJD-PRICELIST.xlsx"
    
    if not Path(file_path).exists():
        print(f"Error: File not found at {file_path}")
        return
    
    print("="*60)
    print("MJD PRICELIST EXTRACTION V2")
    print("With Original Code Preservation & Qwen Enhancement")
    print("="*60)
    
    # Ask about Qwen enhancement
    use_qwen = True
    if not OPENAI_AVAILABLE:
        print("\nWarning: OpenAI library not installed!")
        print("Run: pip install openai")
        response = input("Continue without enhancement? (y/n): ")
        if response.lower() != 'y':
            return
        use_qwen = False
    else:
        response = input("\nUse Qwen 72B enhancement? (y/n): ")
        use_qwen = response.lower() == 'y'
    
    # Create extractor
    extractor = ImprovedPricelistExtractor(file_path, use_qwen=use_qwen)
    
    # Extract
    print("\nStarting extraction...")
    items = extractor.extract_all_sheets()
    
    # Export
    print("\nExporting results...")
    extractor.export_to_csv("pricelist_v2.csv")
    extractor.export_to_json("pricelist_v2.json")
    
    print("\n" + "="*60)
    print("EXTRACTION COMPLETE!")
    print("="*60)
    print("\nOutput files:")
    print("  pricelist_v2.csv - Deduplicated with original codes")
    print("  pricelist_v2.json - Same in JSON format")
    
    # Show samples
    if items:
        print("\nSample items with original codes:")
        samples = [item for item in items if item.original_code][:5]
        for item in samples:
            print(f"\n  ID: {item.id}")
            print(f"  Original Code: {item.original_code}")
            print(f"  Code: {item.code}")
            print(f"  Description: {item.description[:60]}...")
            if item.keywords:
                print(f"  Keywords: {', '.join(item.keywords[:3])}")

if __name__ == "__main__":
    main()