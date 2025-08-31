"""
Final MJD Pricelist Extraction Script with Updated Schema and OpenAI Enhancement
Matches the exact database schema structure
"""

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import numpy as np
import json
import csv
import re
from typing import Dict, List, Optional, Tuple, Any
from dataclasses import dataclass, asdict, field
from pathlib import Path
import openai
import os
from datetime import datetime
import time
import hashlib
import uuid

# OpenAI Configuration - Set your API key as environment variable
OPENAI_API_KEY = os.getenv('OPENAI_API_KEY', '')  # Set OPENAI_API_KEY environment variable

@dataclass
class CellRate:
    """Simplified cell rate reference"""
    reference: str  # e.g. "Groundworks!C14"
    rate: float

@dataclass
class Pattern:
    """Pattern for BOQ matching"""
    boqDescription: str
    boqUnit: Optional[str] = None
    confidence: float = 0.0
    method: str = "initial"
    usageCount: int = 0
    lastUsed: int = 0
    patternId: Optional[str] = None

@dataclass
class PriceItem:
    """Schema matching the database structure exactly"""
    # Core identification fields
    id: str                                    # Unique identifier
    description: str                           # Item description (required)
    
    # Optional fields
    code: Optional[str] = None                # Item code
    ref: Optional[str] = None                 # Reference number
    keywords: Optional[List[str]] = None      # Keywords for searching
    patterns: Optional[List[Dict]] = None     # Learning patterns
    
    # Cell-based rate reference (SIMPLIFIED)
    cellRate: Optional[Dict] = None           # Single cell reference
    
    # Essential categorization fields
    category: Optional[str] = None            # Main category
    subcategory: Optional[str] = None         # Subcategory  
    work_type: Optional[str] = None           # Type of work
    brand: Optional[str] = None               # Brand if applicable
    unit: Optional[str] = None                # Unit of measurement
    rate: Optional[float] = None              # Primary rate
    
    # Excel mapping fields for rate updates
    excelCellReference: Optional[str] = None  # e.g., "Groundworks!F11"
    sourceSheetName: Optional[str] = None     # e.g., "Groundworks"
    sourceRowNumber: Optional[int] = None     # e.g., 11
    sourceColumnLetter: Optional[str] = None  # e.g., "F"
    
    # Legacy fields (keeping for compatibility)
    subCategoryCode: Optional[str] = None
    subCategoryName: Optional[str] = None
    sub_category: Optional[str] = None        # Alternative subcategory field
    
    # Metadata
    isActive: bool = True                     # Whether item is active
    createdAt: int = 0                        # Creation timestamp
    updatedAt: int = 0                        # Last update timestamp
    createdBy: str = "system"                 # User who created the item

class OpenAIEnhancer:
    """Handles OpenAI API calls for data quality improvement"""
    
    def __init__(self, api_key: str = OPENAI_API_KEY):
        self.api_key = api_key
        if api_key:
            self.client = openai.OpenAI(api_key=api_key)
            print("OpenAI client initialized successfully")
        else:
            self.client = None
            print("Warning: No OpenAI API key - enhancement disabled")
        self.cache = {}
    
    def enhance_batch_items(self, items: List[Dict], sheet_name: str) -> List[Dict]:
        """Enhance a batch of items using OpenAI GPT-4"""
        if not self.client:
            return items
        
        # Process in batches of 5 for optimal API usage
        batch_size = 5
        enhanced_items = []
        
        total_batches = (len(items) + batch_size - 1) // batch_size
        
        for i in range(0, len(items), batch_size):
            batch_num = (i // batch_size) + 1
            print(f"    Processing batch {batch_num}/{total_batches}...")
            
            batch = items[i:i + batch_size]
            enhanced_batch = self._process_batch(batch, sheet_name)
            enhanced_items.extend(enhanced_batch)
            
            # Rate limiting
            if batch_num < total_batches:
                time.sleep(1)
        
        return enhanced_items
    
    def _process_batch(self, batch: List[Dict], sheet_name: str) -> List[Dict]:
        """Process a batch of items with OpenAI"""
        try:
            # Prepare batch data
            batch_text = json.dumps([{
                'index': idx,
                'description': item.get('description', ''),
                'unit': item.get('unit', ''),
                'rate': item.get('rate', ''),
                'code': item.get('code', '')
            } for idx, item in enumerate(batch)], indent=2)
            
            prompt = f"""You are a construction cost estimator reviewing pricelist items extracted from Excel.
            
Sheet: {sheet_name}
Items to enhance:
{batch_text}

For each item, improve data quality while PRESERVING original meaning:

1. **Description**: 
   - Complete truncated text based on construction context
   - Fix typos and formatting
   - Keep product codes/standards (BS, EN, etc.)
   - Make professional and clear
   - DO NOT fundamentally change the meaning

2. **Unit**: Standardize to one of:
   - Length: m, lm (linear meter)
   - Area: m², m2
   - Volume: m³, m3
   - Weight: kg, tonnes, t
   - Count: nr, no, each, item
   - Time: hour, hr, day, week, month
   - Lump sum: sum, ls, item
   - Percentage: %

3. **Work Type**: Identify the type of work:
   - Excavation, Concrete, Formwork, Reinforcement, Masonry
   - Drainage, Plumbing, Electrical, Mechanical, HVAC
   - Finishes, Painting, Flooring, Roofing, Insulation
   - External Works, Landscaping, Paving, Fencing
   - Preliminaries, Temporary Works, Site Setup

4. **Subcategory**: Based on {sheet_name} and description, assign appropriate subcategory:
   - Groundworks: Site Clearance, Excavation, Earthworks Support, Disposal, Filling, Foundations
   - RC Works: In-situ Concrete, Formwork, Reinforcement, Precast Concrete, Post-tensioning
   - Drainage: Below Ground Drainage, Above Ground Drainage, Manholes, Pipework, Gullies
   - Services: Electrical Installation, Mechanical Services, Plumbing, HVAC, Fire Systems
   - External Works: Paving, Kerbs, Fencing, Landscaping, Street Furniture, Signage

5. **Keywords**: Generate 4-6 specific search keywords

Return JSON array:
[{{
  "index": 0,
  "description": "enhanced description",
  "unit": "standardized unit",
  "work_type": "type of work",
  "subcategory": "appropriate subcategory",
  "keywords": ["keyword1", "keyword2", "keyword3", "keyword4"],
  "improvements": "what was improved"
}}]

Return ONLY valid JSON array."""

            response = self.client.chat.completions.create(
                model="gpt-3.5-turbo",
                messages=[
                    {"role": "system", "content": "You are an expert construction estimator. Return only valid JSON."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.2,
                max_tokens=2500
            )
            
            # Parse response
            result_text = response.choices[0].message.content.strip()
            
            # Clean JSON response
            if '```json' in result_text:
                result_text = result_text.split('```json')[1].split('```')[0].strip()
            elif '```' in result_text:
                result_text = result_text.split('```')[1].split('```')[0].strip()
            
            # Find JSON array
            import re
            json_match = re.search(r'\[[\s\S]*\]', result_text)
            if json_match:
                result_text = json_match.group()
            
            # Parse JSON
            try:
                enhanced_data = json.loads(result_text)
            except json.JSONDecodeError as e:
                print(f"      JSON parse error: {e}")
                return batch
            
            # Merge enhancements
            enhanced_batch = []
            for item in batch:
                enhanced_item = item.copy()
                
                # Find matching enhancement
                item_index = batch.index(item)
                for enhanced in enhanced_data:
                    if enhanced.get('index') == item_index:
                        # Apply enhancements
                        if enhanced.get('description'):
                            enhanced_item['enhanced_description'] = enhanced['description']
                            enhanced_item['description'] = enhanced['description']
                        
                        if enhanced.get('unit'):
                            enhanced_item['unit'] = enhanced['unit']
                        
                        if enhanced.get('work_type'):
                            enhanced_item['work_type'] = enhanced['work_type']
                        
                        if enhanced.get('subcategory'):
                            enhanced_item['subcategory'] = enhanced['subcategory']
                        
                        if enhanced.get('keywords'):
                            enhanced_item['keywords'] = enhanced['keywords']
                        
                        enhanced_item['quality_enhanced'] = True
                        break
                
                enhanced_batch.append(enhanced_item)
            
            return enhanced_batch
            
        except Exception as e:
            print(f"      Error in OpenAI enhancement: {str(e)[:100]}")
            return batch

class SheetExtractor:
    """Base class for sheet-specific extraction logic"""
    
    def __init__(self, sheet_name: str, worksheet, enhancer: Optional[OpenAIEnhancer] = None):
        self.sheet_name = sheet_name
        self.worksheet = worksheet
        self.enhancer = enhancer
        self.items = []
        self.current_timestamp = int(datetime.now().timestamp() * 1000)
        
    def get_cell_reference(self, row: int, col: int) -> str:
        """Get Excel-style cell reference"""
        return f"{get_column_letter(col)}{row}"
    
    def get_full_cell_reference(self, row: int, col: int) -> str:
        """Get full cell reference with sheet name"""
        return f"{self.sheet_name}!{get_column_letter(col)}{row}"
    
    def clean_value(self, value: Any) -> Any:
        """Clean cell values"""
        if value is None:
            return None
        if isinstance(value, str):
            # Remove excessive whitespace
            cleaned = ' '.join(value.split())
            # Remove non-printable characters
            cleaned = ''.join(char for char in cleaned if char.isprintable() or char in '\n\r\t')
            return cleaned
        return value
    
    def is_valid_rate(self, value: Any) -> bool:
        """Check if value is a valid rate"""
        if value is None:
            return False
        # Skip text values
        if isinstance(value, str):
            # Common non-rate text values
            skip_values = ['by others', 'included', 'n/a', 'nil', '-', 'tbc', 'tbd']
            if any(skip in value.lower() for skip in skip_values):
                return False
        try:
            rate = float(value)
            return 0 < rate < 1000000
        except:
            return False
    
    def is_valid_description(self, desc: Any) -> bool:
        """Check if description is valid"""
        if not desc:
            return False
        
        # Convert to string if needed
        desc = str(desc) if not isinstance(desc, str) else desc
        
        # Minimum length
        if len(desc) < 3:
            return False
        
        # Must have some letters
        if not any(c.isalpha() for c in desc):
            return False
        
        # Skip summary rows
        skip_terms = ['total', 'subtotal', 'carried forward', 'brought forward',
                     'continued', 'see over', 'blank', 'n/a', 'nil']
        
        desc_lower = desc.lower()
        return not any(term in desc_lower for term in skip_terms)
    
    def generate_id(self, code: str) -> str:
        """Generate unique ID"""
        # Use sheet name and code for readable IDs
        base_id = f"{self.sheet_name.replace(' ', '_')}_{code}"
        # Add short hash for uniqueness
        hash_suffix = hashlib.md5(f"{base_id}_{self.current_timestamp}".encode()).hexdigest()[:6]
        return f"{base_id}_{hash_suffix}"
    
    def extract_items(self) -> List[PriceItem]:
        """Main extraction method - must be implemented by subclasses"""
        raise NotImplementedError("Each sheet extractor must implement extract_items")
    
    def post_process_items(self, items: List[PriceItem]) -> List[PriceItem]:
        """Post-process items with OpenAI enhancement"""
        if not self.enhancer or not items:
            return items
        
        print(f"  Enhancing {len(items)} items with OpenAI GPT-4...")
        
        # Convert to dicts for enhancement
        items_dict = []
        for item in items:
            item_dict = {
                'description': item.description,
                'unit': item.unit,
                'subcategory': item.subcategory,
                'rate': item.rate,
                'code': item.code
            }
            items_dict.append(item_dict)
        
        # Enhance with OpenAI
        enhanced_dicts = self.enhancer.enhance_batch_items(items_dict, self.sheet_name)
        
        # Apply enhancements
        for i, item in enumerate(items):
            if i < len(enhanced_dicts):
                enhanced = enhanced_dicts[i]
                
                # Apply enhancements
                if enhanced.get('description'):
                    item.description = enhanced['description']
                
                if enhanced.get('unit'):
                    item.unit = enhanced['unit']
                
                if enhanced.get('work_type'):
                    item.work_type = enhanced['work_type']
                
                if enhanced.get('subcategory'):
                    item.subcategory = enhanced['subcategory']
                    # Also set legacy fields
                    item.sub_category = enhanced['subcategory']
                    item.subCategoryName = enhanced['subcategory']
                
                if enhanced.get('keywords'):
                    item.keywords = enhanced['keywords']
        
        return items

class GroundworksExtractor(SheetExtractor):
    """Extractor for Groundworks sheet"""
    
    def extract_items(self) -> List[PriceItem]:
        items = []
        current_subcategory = None
        item_counter = 1
        
        # Find data start row
        data_start_row = 10
        for row in range(1, min(30, self.worksheet.max_row + 1)):
            cell = self.worksheet.cell(row=row, column=2)
            if cell.value and 'Description' in str(cell.value):
                data_start_row = row + 2
                break
        
        # Process rows
        for row in range(data_start_row, self.worksheet.max_row + 1):
            cell_b = self.worksheet.cell(row=row, column=2)
            
            # Check for subcategory headers
            if cell_b.value:
                cell_value = self.clean_value(cell_b.value)
                
                # Check if this is a subcategory
                is_subcategory = False
                if cell_b.font and cell_b.font.bold:
                    is_subcategory = True
                
                if cell_value and cell_value.isupper() and len(cell_value) < 50:
                    is_subcategory = True
                
                subcategory_keywords = ['excavat', 'disposal', 'filling', 'earthwork',
                                       'foundation', 'drainage', 'concrete', 'piling']
                if any(keyword in (cell_value or '').lower() for keyword in subcategory_keywords):
                    is_subcategory = True
                
                if is_subcategory:
                    current_subcategory = cell_value
                    continue
            
            # Extract item data
            description = self.clean_value(cell_b.value)
            
            if not self.is_valid_description(description):
                continue
            
            # Get item code from column A
            item_code = self.clean_value(self.worksheet.cell(row=row, column=1).value)
            
            # Get data
            quantity = self.worksheet.cell(row=row, column=4).value
            unit = self.clean_value(self.worksheet.cell(row=row, column=5).value)
            rate = self.worksheet.cell(row=row, column=6).value
            
            # Find valid rate
            rate_col = 6
            if not self.is_valid_rate(rate):
                for col in [7, 8, 9]:
                    alt_rate = self.worksheet.cell(row=row, column=col).value
                    if self.is_valid_rate(alt_rate):
                        rate = alt_rate
                        rate_col = col
                        break
            
            # Process valid items
            if description and (rate or quantity):
                # Generate code
                if item_code and str(item_code).strip():
                    code = str(item_code).strip()
                else:
                    code = f"GW{item_counter:04d}"
                
                # Generate unique ID
                item_id = self.generate_id(code)
                
                # Create cell rate reference
                cell_rate = None
                if self.is_valid_rate(rate):
                    cell_rate = {
                        'reference': self.get_full_cell_reference(row, rate_col),
                        'rate': float(rate)
                    }
                
                # Auto-detect subcategory if needed
                if not current_subcategory:
                    desc_lower = description.lower()
                    if 'excavat' in desc_lower:
                        current_subcategory = 'Excavation'
                    elif 'fill' in desc_lower:
                        current_subcategory = 'Filling'
                    elif 'disposal' in desc_lower or 'cart away' in desc_lower:
                        current_subcategory = 'Disposal'
                    elif 'clear' in desc_lower or 'demolish' in desc_lower:
                        current_subcategory = 'Site Clearance'
                    else:
                        current_subcategory = 'General Groundworks'
                
                # Create item
                item = PriceItem(
                    id=item_id,
                    code=code,
                    description=description,
                    unit=unit,
                    rate=float(rate) if self.is_valid_rate(rate) else None,
                    category=self.sheet_name,
                    subcategory=current_subcategory,
                    sub_category=current_subcategory,  # Legacy field
                    cellRate=cell_rate,
                    excelCellReference=self.get_full_cell_reference(row, rate_col) if rate else None,
                    sourceSheetName=self.sheet_name,
                    sourceRowNumber=row,
                    sourceColumnLetter=get_column_letter(rate_col) if rate else None,
                    isActive=True,
                    createdAt=self.current_timestamp,
                    updatedAt=self.current_timestamp,
                    createdBy="system"
                )
                
                items.append(item)
                item_counter += 1
        
        return self.post_process_items(items)

class RCWorksExtractor(SheetExtractor):
    """Extractor for RC Works sheet"""
    
    def extract_items(self) -> List[PriceItem]:
        items = []
        current_subcategory = None
        item_counter = 1
        
        # Find data start
        data_start_row = 10
        for row in range(1, min(30, self.worksheet.max_row + 1)):
            cell = self.worksheet.cell(row=row, column=2)
            if cell.value and 'Description' in str(cell.value):
                data_start_row = row + 2
                break
        
        for row in range(data_start_row, self.worksheet.max_row + 1):
            cell_b = self.worksheet.cell(row=row, column=2)
            
            # Check for subcategory headers
            if cell_b.value:
                value_str = self.clean_value(cell_b.value)
                
                # RC specific subcategories
                rc_categories = {
                    'in-situ concrete': 'In-situ Concrete',
                    'in situ concrete': 'In-situ Concrete',
                    'formwork': 'Formwork',
                    'reinforcement': 'Reinforcement',
                    'rebar': 'Reinforcement',
                    'precast': 'Precast Concrete',
                    'post-tension': 'Post-tensioning',
                    'prestress': 'Post-tensioning',
                    'sundries': 'Concrete Sundries',
                    'accessories': 'Concrete Accessories'
                }
                
                value_lower = (value_str or '').lower()
                for key, category in rc_categories.items():
                    if key in value_lower:
                        current_subcategory = category
                        continue
            
            description = self.clean_value(cell_b.value)
            
            if not self.is_valid_description(description):
                continue
            
            # Get data
            item_code = self.clean_value(self.worksheet.cell(row=row, column=1).value)
            quantity = self.worksheet.cell(row=row, column=4).value
            unit = self.clean_value(self.worksheet.cell(row=row, column=5).value)
            rate = self.worksheet.cell(row=row, column=6).value
            rate_col = 6
            
            if description and (rate or quantity):
                # Generate code
                if item_code and str(item_code).strip():
                    code = str(item_code).strip()
                else:
                    code = f"RC{item_counter:04d}"
                
                item_id = self.generate_id(code)
                
                # Create cell rate
                cell_rate = None
                if self.is_valid_rate(rate):
                    cell_rate = {
                        'reference': self.get_full_cell_reference(row, rate_col),
                        'rate': float(rate)
                    }
                
                # Auto-detect subcategory
                if not current_subcategory:
                    desc_lower = description.lower()
                    if 'concrete' in desc_lower and 'formwork' not in desc_lower:
                        current_subcategory = 'In-situ Concrete'
                    elif 'formwork' in desc_lower or 'shutter' in desc_lower:
                        current_subcategory = 'Formwork'
                    elif 'reinforc' in desc_lower or 'rebar' in desc_lower or 'mesh' in desc_lower:
                        current_subcategory = 'Reinforcement'
                    else:
                        current_subcategory = 'General RC Works'
                
                item = PriceItem(
                    id=item_id,
                    code=code,
                    description=description,
                    unit=unit,
                    rate=float(rate) if self.is_valid_rate(rate) else None,
                    category=self.sheet_name,
                    subcategory=current_subcategory,
                    sub_category=current_subcategory,
                    cellRate=cell_rate,
                    excelCellReference=self.get_full_cell_reference(row, rate_col) if rate else None,
                    sourceSheetName=self.sheet_name,
                    sourceRowNumber=row,
                    sourceColumnLetter=get_column_letter(rate_col) if rate else None,
                    isActive=True,
                    createdAt=self.current_timestamp,
                    updatedAt=self.current_timestamp,
                    createdBy="system"
                )
                
                items.append(item)
                item_counter += 1
        
        return self.post_process_items(items)

class GenericExtractor(SheetExtractor):
    """Generic extractor for all other sheets"""
    
    def extract_items(self) -> List[PriceItem]:
        items = []
        current_subcategory = None
        item_counter = 1
        
        # Auto-detect columns
        desc_col = 2
        rate_col = 6
        unit_col = 5
        code_col = 1
        
        # Find headers
        for row in range(1, min(20, self.worksheet.max_row + 1)):
            for col in range(1, min(15, self.worksheet.max_column + 1)):
                cell_value = str(self.worksheet.cell(row=row, column=col).value or '').lower()
                if 'description' in cell_value:
                    desc_col = col
                elif 'rate' in cell_value:
                    rate_col = col
                elif 'unit' in cell_value:
                    unit_col = col
                elif 'code' in cell_value or 'ref' in cell_value:
                    code_col = col
        
        # Start extraction
        data_start_row = 10
        
        for row in range(data_start_row, self.worksheet.max_row + 1):
            description = self.clean_value(self.worksheet.cell(row=row, column=desc_col).value)
            
            if not self.is_valid_description(description):
                continue
            
            item_code = self.clean_value(self.worksheet.cell(row=row, column=code_col).value)
            unit = self.clean_value(self.worksheet.cell(row=row, column=unit_col).value)
            rate = self.worksheet.cell(row=row, column=rate_col).value
            
            # Find valid rate in nearby columns
            if not self.is_valid_rate(rate):
                for col in range(max(1, rate_col - 2), min(self.worksheet.max_column + 1, rate_col + 3)):
                    alt_rate = self.worksheet.cell(row=row, column=col).value
                    if self.is_valid_rate(alt_rate):
                        rate = alt_rate
                        rate_col = col
                        break
            
            if description and self.is_valid_rate(rate):
                # Generate code
                if item_code and str(item_code).strip():
                    code = str(item_code).strip()
                else:
                    # Create prefix from sheet name
                    prefix = ''.join([c for c in self.sheet_name[:3].upper() if c.isalpha()])[:2]
                    if not prefix:
                        prefix = 'GN'
                    code = f"{prefix}{item_counter:04d}"
                
                item_id = self.generate_id(code)
                
                # Create cell rate
                cell_rate = {
                    'reference': self.get_full_cell_reference(row, rate_col),
                    'rate': float(rate)
                }
                
                # Infer subcategory from sheet name and description
                if not current_subcategory:
                    current_subcategory = f"General {self.sheet_name}"
                
                item = PriceItem(
                    id=item_id,
                    code=code,
                    description=description,
                    unit=unit,
                    rate=float(rate) if self.is_valid_rate(rate) else None,
                    category=self.sheet_name,
                    subcategory=current_subcategory,
                    sub_category=current_subcategory,
                    cellRate=cell_rate,
                    excelCellReference=self.get_full_cell_reference(row, rate_col),
                    sourceSheetName=self.sheet_name,
                    sourceRowNumber=row,
                    sourceColumnLetter=get_column_letter(rate_col),
                    isActive=True,
                    createdAt=self.current_timestamp,
                    updatedAt=self.current_timestamp,
                    createdBy="system"
                )
                
                items.append(item)
                item_counter += 1
        
        return self.post_process_items(items)

class FinalPricelistExtractor:
    """Main extractor with all sheets and OpenAI enhancement"""
    
    SHEET_EXTRACTORS = {
        'Groundworks': GroundworksExtractor,
        'RC works': RCWorksExtractor,
        'RC Works': RCWorksExtractor,  # Alternative naming
        # All other sheets use GenericExtractor
    }
    
    SKIP_SHEETS = ['Summary', 'Set factors & prices', 'Tender Summary', 'Budget Costings']
    
    def __init__(self, file_path: str, use_openai: bool = True):
        self.file_path = file_path
        self.use_openai = use_openai
        self.workbook = None
        self.all_items = []
        
        # Initialize OpenAI enhancer
        if use_openai:
            self.enhancer = OpenAIEnhancer(OPENAI_API_KEY)
        else:
            self.enhancer = None
    
    def extract_all_sheets(self) -> List[PriceItem]:
        """Extract items from all sheets"""
        self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
        
        print(f"\nFound {len(self.workbook.sheetnames)} sheets in workbook")
        print("="*60)
        
        for sheet_name in self.workbook.sheetnames:
            if sheet_name in self.SKIP_SHEETS:
                print(f"Skipping: {sheet_name}")
                continue
            
            print(f"\nProcessing: {sheet_name}")
            worksheet = self.workbook[sheet_name]
            
            # Select appropriate extractor
            if sheet_name in self.SHEET_EXTRACTORS:
                extractor_class = self.SHEET_EXTRACTORS[sheet_name]
            else:
                extractor_class = GenericExtractor
            
            # Create extractor
            extractor = extractor_class(sheet_name, worksheet, self.enhancer)
            
            # Extract items
            items = extractor.extract_items()
            
            if items:
                self.all_items.extend(items)
                print(f"  Extracted {len(items)} items")
            else:
                print(f"  No items found")
        
        self.workbook.close()
        
        # Print summary
        self.print_extraction_summary()
        
        return self.all_items
    
    def print_extraction_summary(self):
        """Print extraction summary"""
        print("\n" + "="*60)
        print("EXTRACTION SUMMARY")
        print("="*60)
        print(f"Total items extracted: {len(self.all_items)}")
        
        # Count by category
        categories = {}
        for item in self.all_items:
            cat = item.category or "Uncategorized"
            categories[cat] = categories.get(cat, 0) + 1
        
        print("\nItems by category:")
        for cat, count in sorted(categories.items(), key=lambda x: x[1], reverse=True):
            print(f"  {cat}: {count}")
        
        # Quality metrics
        with_keywords = sum(1 for item in self.all_items if item.keywords)
        with_subcategory = sum(1 for item in self.all_items if item.subcategory)
        with_work_type = sum(1 for item in self.all_items if item.work_type)
        
        print(f"\nQuality metrics:")
        print(f"  With keywords: {with_keywords}/{len(self.all_items)}")
        print(f"  With subcategory: {with_subcategory}/{len(self.all_items)}")
        print(f"  With work type: {with_work_type}/{len(self.all_items)}")
    
    def export_to_csv(self, output_file: str = "pricelist_final.csv"):
        """Export to CSV with full schema"""
        if not self.all_items:
            print("No items to export")
            return
        
        csv_data = []
        for item in self.all_items:
            row = {
                # Core fields
                'id': item.id,
                'code': item.code or '',
                'ref': item.ref or '',
                'description': item.description,
                'keywords': '|'.join(item.keywords) if item.keywords else '',
                
                # Categorization
                'category': item.category or '',
                'subcategory': item.subcategory or '',
                'work_type': item.work_type or '',
                'brand': item.brand or '',
                'unit': item.unit or '',
                'rate': item.rate or '',
                
                # Cell reference (simplified)
                'cellRate_reference': item.cellRate['reference'] if item.cellRate else '',
                'cellRate_rate': item.cellRate['rate'] if item.cellRate else '',
                
                # Excel mapping
                'excelCellReference': item.excelCellReference or '',
                'sourceSheetName': item.sourceSheetName or '',
                'sourceRowNumber': item.sourceRowNumber or '',
                'sourceColumnLetter': item.sourceColumnLetter or '',
                
                # Metadata
                'isActive': item.isActive,
                'createdAt': item.createdAt,
                'updatedAt': item.updatedAt,
                'createdBy': item.createdBy
            }
            
            csv_data.append(row)
        
        # Write CSV
        if csv_data:
            fieldnames = list(csv_data[0].keys())
            
            with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(csv_data)
            
            print(f"\nExported {len(csv_data)} items to {output_file}")
    
    def export_to_json(self, output_file: str = "pricelist_final.json"):
        """Export to JSON with full schema"""
        if not self.all_items:
            print("No items to export")
            return
        
        json_data = []
        for item in self.all_items:
            item_dict = {
                # Core identification
                'id': item.id,
                'code': item.code,
                'ref': item.ref,
                'description': item.description,
                'keywords': item.keywords,
                'patterns': item.patterns or [],
                
                # Cell rate (simplified)
                'cellRate': item.cellRate,
                
                # Categorization
                'category': item.category,
                'subcategory': item.subcategory,
                'work_type': item.work_type,
                'brand': item.brand,
                'unit': item.unit,
                'rate': item.rate,
                
                # Excel mapping
                'excelCellReference': item.excelCellReference,
                'sourceSheetName': item.sourceSheetName,
                'sourceRowNumber': item.sourceRowNumber,
                'sourceColumnLetter': item.sourceColumnLetter,
                
                # Legacy fields
                'subCategoryCode': item.subCategoryCode,
                'subCategoryName': item.subCategoryName,
                'sub_category': item.sub_category,
                
                # Metadata
                'isActive': item.isActive,
                'createdAt': item.createdAt,
                'updatedAt': item.updatedAt,
                'createdBy': item.createdBy
            }
            
            json_data.append(item_dict)
        
        # Write JSON
        with open(output_file, 'w', encoding='utf-8') as jsonfile:
            json.dump(json_data, jsonfile, indent=2, ensure_ascii=False)
        
        print(f"Exported {len(json_data)} items to {output_file}")

def main():
    """Main execution"""
    file_path = r"C:\Users\abaza\pricelist extraction\MJD-PRICELIST.xlsx"
    
    if not Path(file_path).exists():
        print(f"Error: File not found at {file_path}")
        return
    
    print("="*60)
    print("MJD PRICELIST EXTRACTION - FINAL VERSION")
    print("="*60)
    print(f"Input file: {file_path}")
    print(f"OpenAI Enhancement: Enabled")
    print(f"Schema: Updated to match database structure")
    
    # Create extractor
    extractor = FinalPricelistExtractor(file_path, use_openai=True)
    
    # Extract all sheets
    print("\nStarting extraction with OpenAI enhancement...")
    items = extractor.extract_all_sheets()
    
    # Export results
    print("\n" + "="*60)
    print("EXPORTING RESULTS")
    print("="*60)
    
    extractor.export_to_csv("pricelist_final.csv")
    extractor.export_to_json("pricelist_final.json")
    
    print("\n" + "="*60)
    print("EXTRACTION COMPLETE!")
    print("="*60)
    print("\nOutput files:")
    print("  pricelist_final.csv - For spreadsheet import")
    print("  pricelist_final.json - For database import")
    
    # Show sample
    if items:
        print("\nSample items:")
        for item in items[:3]:
            print(f"\n  ID: {item.id}")
            print(f"  Code: {item.code}")
            print(f"  Description: {item.description[:60]}...")
            print(f"  Category: {item.category} > {item.subcategory}")
            if item.keywords:
                print(f"  Keywords: {', '.join(item.keywords[:3])}")
            if item.cellRate:
                print(f"  Cell Ref: {item.cellRate['reference']} = {item.cellRate['rate']}")

if __name__ == "__main__":
    main()