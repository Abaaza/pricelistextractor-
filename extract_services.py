"""
Extraction script for Services sheet
Handles the specific structure and format of the Services works pricelist
"""

import pandas as pd
import numpy as np
import json
import re
from datetime import datetime
from pathlib import Path
import string

class ServicesExtractor:
    def __init__(self, excel_file='MJD-PRICELIST.xlsx'):
        self.excel_file = excel_file
        self.sheet_name = 'Services'
        self.df = None
        self.extracted_items = []
        
    def load_sheet(self):
        """Load the Services sheet"""
        print(f"Loading {self.sheet_name} sheet...")
        self.df = pd.read_excel(self.excel_file, sheet_name=self.sheet_name, header=None)
        print(f"Loaded {len(self.df)} rows x {len(self.df.columns)} columns")
        return self.df
    
    def identify_data_rows(self):
        """Identify rows containing actual pricelist data"""
        data_rows = []
        
        for idx, row in self.df.iterrows():
            # Skip if row is mostly empty
            if row.notna().sum() < 3:
                continue
                
            # Look for patterns that indicate services data
            code_col = row[0] if 0 < len(row) else None
            if pd.notna(code_col):
                code_str = str(code_col).strip()
                # Services codes often start with S, SV, M&E, or numbers
                if (re.match(r'^\d+', code_str) or 
                    re.match(r'^S\d+', code_str, re.I) or
                    re.match(r'^SV', code_str, re.I) or
                    re.match(r'^M&E', code_str, re.I) or
                    re.match(r'^[A-Z]+\d+', code_str)):
                    data_rows.append(idx)
                    continue
            
            # Check if row has services-related content
            for col_idx in range(1, min(5, len(row))):
                cell = row[col_idx]
                if pd.notna(cell):
                    cell_str = str(cell).strip().lower()
                    # Services keywords
                    if any(keyword in cell_str for keyword in 
                           ['electrical', 'plumbing', 'hvac', 'mechanical', 'cable',
                            'conduit', 'wire', 'socket', 'switch', 'light', 'power',
                            'distribution', 'panel', 'breaker', 'transformer',
                            'water supply', 'hot water', 'cold water', 'gas',
                            'ventilation', 'air conditioning', 'heating', 'boiler',
                            'pump', 'valve', 'meter', 'sensor', 'control',
                            'fire alarm', 'sprinkler', 'detection', 'emergency',
                            'data', 'communication', 'network', 'telephone']):
                        data_rows.append(idx)
                        break
        
        return data_rows
    
    def extract_code(self, row, col_idx=0):
        """Extract code from row"""
        if col_idx < len(row) and pd.notna(row[col_idx]):
            code = str(row[col_idx]).strip()
            # Clean up code
            code = re.sub(r'\s+', '', code)
            if code and not code.lower() in ['nan', 'none', '-', '']:
                return code
        return None
    
    def extract_description(self, row, start_col=1):
        """Extract and clean description for services"""
        description_parts = []
        
        # Collect description from multiple columns
        for col_idx in range(start_col, min(start_col + 3, len(row))):
            if pd.notna(row[col_idx]):
                part = str(row[col_idx]).strip()
                # Skip if it's a number or unit
                if not re.match(r'^[\d,\.]+$', part) and not self.is_unit(part):
                    description_parts.append(part)
        
        description = ' '.join(description_parts)
        
        # Clean and expand services-specific abbreviations
        replacements = {
            ' elec ': ' electrical ',
            ' mech ': ' mechanical ',
            ' hvac ': ' HVAC ',
            ' a/c ': ' air conditioning ',
            ' ac ': ' air conditioning ',
            ' hw ': ' hot water ',
            ' cw ': ' cold water ',
            ' lwc ': ' low water content ',
            ' swc ': ' soil and waste ',
            ' rwp ': ' rainwater pipe ',
            ' svp ': ' soil vent pipe ',
            ' dia ': ' diameter ',
            ' thk ': ' thick ',
            ' galv ': ' galvanized ',
            ' gi ': ' galvanized iron ',
            ' ms ': ' mild steel ',
            ' ss ': ' stainless steel ',
            ' cu ': ' copper ',
            ' pvc ': ' PVC ',
            ' cpvc ': ' CPVC ',
            ' ppr ': ' PPR ',
            ' hdpe ': ' HDPE ',
            ' incl ': ' including ',
            ' excl ': ' excluding ',
            ' c/w ': ' complete with ',
            ' w/ ': ' with ',
            ' w/o ': ' without ',
            ' db ': ' distribution board ',
            ' mcb ': ' miniature circuit breaker ',
            ' mccb ': ' molded case circuit breaker ',
            ' rccb ': ' residual current circuit breaker ',
            ' fcu ': ' fan coil unit ',
            ' ahu ': ' air handling unit ',
            ' vrf ': ' variable refrigerant flow ',
            ' btu ': ' BTU ',
            ' tr ': ' ton refrigeration ',
            ' kw ': ' kilowatt ',
            ' hp ': ' horsepower ',
            ' lux ': ' lux ',
            ' ip ': ' IP rating ',
            ' cat ': ' category ',
            ' swa ': ' steel wire armored ',
            ' xlpe ': ' XLPE ',
            ' lv ': ' low voltage ',
            ' mv ': ' medium voltage ',
            ' hv ': ' high voltage ',
        }
        
        for old, new in replacements.items():
            description = description.replace(old, new)
            description = description.replace(old.upper(), new)
        
        # Fix patterns
        description = re.sub(r'(\d+)mm\s*dia', r'\1mm diameter', description)
        description = re.sub(r'(\d+)dia', r'\1mm diameter', description)
        description = re.sub(r'(\d+)thk', r'\1mm thick', description)
        description = re.sub(r'(\d+)sqmm', r'\1 sq.mm', description)
        description = re.sub(r'(\d+)c', r'\1 core', description)
        description = re.sub(r'(\d+)\s*[xX]\s*(\d+)', r'\1x\2', description)
        
        # Clean up spaces
        description = ' '.join(description.split())
        
        return description
    
    def is_unit(self, value):
        """Check if value is a unit"""
        if pd.isna(value):
            return False
        
        value_str = str(value).strip().lower()
        
        units = ['m', 'm2', 'm²', 'm3', 'm³', 'nr', 'no', 'item', 'sum',
                 'kg', 'tonnes', 't', 'lm', 'sqm', 'cum', 'each', 'set',
                 'point', 'kw', 'kva', 'amp', 'ton']
        
        return value_str in units
    
    def extract_unit(self, row, expected_col=None):
        """Extract unit from row"""
        # Try expected column first
        if expected_col is not None and expected_col < len(row):
            if pd.notna(row[expected_col]):
                value = str(row[expected_col]).strip()
                if self.is_unit(value):
                    return self.standardize_unit(value)
        
        # Search for unit in other columns
        for col_idx in range(2, min(6, len(row))):
            if pd.notna(row[col_idx]):
                value = str(row[col_idx]).strip()
                if self.is_unit(value):
                    return self.standardize_unit(value)
        
        # Infer from description
        return self.infer_unit_from_description(row)
    
    def standardize_unit(self, unit):
        """Standardize unit format"""
        unit_map = {
            'm2': 'm²', 'sqm': 'm²', 'sq.m': 'm²',
            'm3': 'm³', 'cum': 'm³', 'cu.m': 'm³',
            'no': 'nr', 'no.': 'nr', 'each': 'nr',
            't': 'tonnes', 'ton': 'tonnes', 'tonne': 'tonnes',
            'lm': 'm', 'lin.m': 'm', 'l.m': 'm',
            'pt': 'point', 'pts': 'point',
        }
        
        unit_lower = unit.lower()
        return unit_map.get(unit_lower, unit_lower)
    
    def infer_unit_from_description(self, row):
        """Infer unit from description content for services"""
        desc = self.extract_description(row)
        desc_lower = desc.lower()
        
        # Services specific patterns
        if any(word in desc_lower for word in ['cable', 'wire', 'conduit', 'pipe', 'duct']):
            if 'install' in desc_lower or 'laying' in desc_lower or 'run' in desc_lower:
                return 'm'
            elif 'connection' in desc_lower or 'termination' in desc_lower:
                return 'nr'
            return 'm'
        elif any(word in desc_lower for word in ['socket', 'switch', 'outlet', 'point', 'breaker']):
            return 'point'
        elif any(word in desc_lower for word in ['panel', 'board', 'unit', 'pump', 'motor']):
            return 'nr'
        elif any(word in desc_lower for word in ['light', 'luminaire', 'fixture', 'fitting']):
            return 'nr'
        elif 'testing' in desc_lower or 'commissioning' in desc_lower:
            return 'sum'
        elif any(word in desc_lower for word in ['excavation', 'trench']):
            return 'm³'
        elif 'insulation' in desc_lower:
            if 'pipe' in desc_lower:
                return 'm'
            return 'm²'
        elif any(word in desc_lower for word in ['valve', 'meter', 'sensor', 'detector']):
            return 'nr'
        elif any(word in desc_lower for word in ['kw', 'kilowatt', 'kva']):
            return 'kw'
        elif 'ton' in desc_lower and 'refrigeration' in desc_lower:
            return 'ton'
        
        return 'item'
    
    def extract_rate(self, row, start_col=3):
        """Extract rate value"""
        for col_idx in range(start_col, min(start_col + 4, len(row))):
            if pd.notna(row[col_idx]):
                value = str(row[col_idx]).strip()
                # Check if it's a number
                value_clean = value.replace(',', '').replace('£', '').replace('$', '')
                try:
                    rate = float(value_clean)
                    if rate > 0:  # Valid rate
                        return rate
                except:
                    continue
        return None
    
    def get_cell_reference(self, row_idx, col_idx):
        """Convert row and column index to Excel cell reference"""
        if col_idx < 26:
            col_letter = string.ascii_uppercase[col_idx]
        else:
            col_letter = string.ascii_uppercase[col_idx // 26 - 1] + string.ascii_uppercase[col_idx % 26]
        
        return f"{col_letter}{row_idx + 1}"
    
    def determine_subcategory(self, description):
        """Determine subcategory based on description for services"""
        desc_lower = description.lower()
        
        # Services subcategories
        if any(word in desc_lower for word in ['electrical', 'power', 'cable', 'wire']):
            if 'distribution' in desc_lower or 'panel' in desc_lower or 'board' in desc_lower:
                return 'Electrical Distribution'
            elif 'cable' in desc_lower or 'wire' in desc_lower:
                return 'Cables and Wiring'
            elif 'light' in desc_lower or 'luminaire' in desc_lower:
                return 'Lighting'
            elif 'socket' in desc_lower or 'switch' in desc_lower or 'outlet' in desc_lower:
                return 'Wiring Devices'
            elif 'earthing' in desc_lower or 'grounding' in desc_lower:
                return 'Earthing and Grounding'
            else:
                return 'Electrical Works'
        elif any(word in desc_lower for word in ['plumbing', 'water', 'pipe']):
            if 'hot water' in desc_lower:
                return 'Hot Water System'
            elif 'cold water' in desc_lower or 'potable' in desc_lower:
                return 'Cold Water System'
            elif 'waste' in desc_lower or 'soil' in desc_lower:
                return 'Soil and Waste'
            elif 'sanitary' in desc_lower or 'fixture' in desc_lower:
                return 'Sanitary Fixtures'
            else:
                return 'Plumbing Works'
        elif any(word in desc_lower for word in ['hvac', 'air conditioning', 'ventilation', 'heating']):
            if 'air conditioning' in desc_lower or 'cooling' in desc_lower:
                return 'Air Conditioning'
            elif 'ventilation' in desc_lower or 'exhaust' in desc_lower:
                return 'Ventilation System'
            elif 'heating' in desc_lower or 'boiler' in desc_lower:
                return 'Heating System'
            elif 'duct' in desc_lower:
                return 'Ductwork'
            else:
                return 'HVAC Works'
        elif 'fire' in desc_lower:
            if 'alarm' in desc_lower or 'detection' in desc_lower:
                return 'Fire Alarm System'
            elif 'sprinkler' in desc_lower or 'suppression' in desc_lower:
                return 'Fire Fighting System'
            elif 'extinguisher' in desc_lower:
                return 'Fire Extinguishers'
            else:
                return 'Fire Protection'
        elif any(word in desc_lower for word in ['data', 'network', 'communication', 'telephone']):
            if 'structured cabling' in desc_lower:
                return 'Structured Cabling'
            elif 'network' in desc_lower:
                return 'Network Infrastructure'
            else:
                return 'Low Current Systems'
        elif 'gas' in desc_lower:
            return 'Gas System'
        elif 'lift' in desc_lower or 'elevator' in desc_lower:
            return 'Vertical Transportation'
        elif 'bms' in desc_lower or 'building management' in desc_lower:
            return 'Building Management System'
        elif 'testing' in desc_lower or 'commissioning' in desc_lower:
            return 'Testing and Commissioning'
        else:
            return 'General Services'
    
    def determine_work_type(self, description, subcategory):
        """Determine work type for services"""
        desc_lower = description.lower()
        
        if 'install' in desc_lower:
            return 'Installation'
        elif 'supply' in desc_lower and 'install' in desc_lower:
            return 'Supply and Installation'
        elif 'supply' in desc_lower:
            return 'Supply Only'
        elif 'testing' in desc_lower:
            return 'Testing'
        elif 'commissioning' in desc_lower:
            return 'Commissioning'
        elif 'connection' in desc_lower or 'termination' in desc_lower:
            return 'Connection'
        elif 'excavation' in desc_lower:
            return 'Excavation'
        elif 'maintenance' in desc_lower:
            return 'Maintenance'
        else:
            return 'Services Works'
    
    def generate_keywords(self, description, subcategory):
        """Generate search keywords for services"""
        keywords = []
        desc_lower = description.lower()
        
        # Extract cable sizes
        cable_sizes = re.findall(r'(\d+)\s*(?:x\s*)?(\d+)\s*(?:sq\.?mm|mm2)', desc_lower)
        for size in cable_sizes[:1]:
            keywords.append(f"{size[0]}x{size[1]}sqmm" if size[1] else f"{size[0]}sqmm")
        
        # Extract pipe sizes
        pipe_sizes = re.findall(r'(\d+)mm\s*(?:diameter|dia)', desc_lower)
        for size in pipe_sizes[:1]:
            keywords.append(f"{size}mm")
        
        # Extract power ratings
        power = re.findall(r'(\d+(?:\.\d+)?)\s*(?:kw|kva|hp|amp)', desc_lower)
        for p in power[:1]:
            keywords.append(f"{p}kw")
        
        # Material keywords
        materials = ['copper', 'pvc', 'cpvc', 'hdpe', 'galvanized', 'steel',
                    'aluminum', 'xlpe', 'swa', 'armored']
        for mat in materials:
            if mat in desc_lower:
                keywords.append(mat)
        
        # Key services terms
        terms = ['electrical', 'plumbing', 'hvac', 'cable', 'pipe', 'conduit',
                 'panel', 'socket', 'switch', 'light', 'pump', 'valve',
                 'air_conditioning', 'ventilation', 'fire_alarm']
        
        for term in terms:
            if term.replace('_', ' ') in desc_lower:
                keywords.append(term)
        
        # Add subcategory keyword
        if subcategory:
            keywords.append(subcategory.lower().replace(' ', '_'))
        
        # Limit and remove duplicates
        seen = set()
        unique_keywords = []
        for kw in keywords:
            if kw not in seen:
                seen.add(kw)
                unique_keywords.append(kw)
        
        return unique_keywords[:6]
    
    def is_row_bold(self, row_idx):
        """Check if column B in a row is bold"""
        try:
            from openpyxl import load_workbook
            if not hasattr(self, 'worksheet'):
                wb = load_workbook(self.excel_file, read_only=True, data_only=True)
                self.worksheet = wb[self.sheet_name]
            
            cell = self.worksheet.cell(row=row_idx + 1, column=2)  # Column B
            if cell and cell.font and cell.font.bold:
                return True
            return False
        except:
            return False
    
    def extract_items(self):
        """Main extraction method"""
        if self.df is None:
            self.load_sheet()
        
        print(f"\nExtracting items from {self.sheet_name}...")
        
        # Load worksheet for bold detection
        try:
            from openpyxl import load_workbook
            wb = load_workbook(self.excel_file, read_only=True, data_only=True)
            self.worksheet = wb[self.sheet_name]
        except:
            self.worksheet = None
        
        items = []
        current_id = 1
        current_header = None  # Track current header description
        current_subcategory = "General Services"  # Default subcategory
        
        # Process all rows starting from row 10
        for row_idx in range(10, len(self.df)):
            row = self.df.iloc[row_idx]
            
            # Determine which section we're in and apply appropriate logic
            if 12 <= row_idx <= 307:  # Rows 13-308 (0-indexed: 12-307)
                # Section 1: Header + Range pattern
                # Check if this row has a header description in column B
                if pd.notna(row[1]):
                    desc_text = str(row[1]).strip()
                    # Check if it's a header (has text but no code in column A)
                    if pd.isna(row[0]) and len(desc_text) > 10:
                        # Skip certain rows
                        if not any(skip in desc_text.lower() for skip in ['description', 'preambles', 'all items below']):
                            current_header = desc_text
                            continue
            
            elif 309 <= row_idx <= 345:  # Rows 310-346 (0-indexed: 309-345)
                # Section 2: Normal pattern with bold subcategories
                # Reset header when entering this section
                if row_idx == 309:
                    current_header = None
                
                # Check for bold subcategory
                if pd.notna(row[1]) and self.is_row_bold(row_idx):
                    current_subcategory = str(row[1]).strip()
                    continue
                
                # Special handling for row 346 - set subcategory if not already set
                if row_idx == 345:  # Row 346 (0-indexed)
                    if current_subcategory == "General Services":
                        current_subcategory = "Services Works"
            
            elif 347 <= row_idx <= 687:  # Rows 348-688 (0-indexed: 347-687)
                # Section 3: Header + column C pattern
                # Check if this is a header row (bold text in column B)
                if pd.notna(row[1]) and self.is_row_bold(row_idx):
                    current_header = str(row[1]).strip()
                    continue
            
            elif 689 <= row_idx <= 806:  # Rows 690-807 (0-indexed: 689-806)
                # Section 4: Normal pattern with bold subcategories
                # Check for bold subcategory
                if pd.notna(row[1]) and self.is_row_bold(row_idx):
                    current_subcategory = str(row[1]).strip()
                    continue
            
            # Check if this row has a code in column A
            code = self.extract_code(row)
            if not code:
                continue
            
            # Build description based on section
            description = None
            
            if 12 <= row_idx <= 307:  # Section 1: Header + Range
                # Check if this is a range-based row (has data in columns C-E)
                has_range = (pd.notna(row[2]) or pd.notna(row[3]) or pd.notna(row[4]))
                
                if has_range and current_header:
                    # This is a range row - combine header with range info
                    # Collect non-empty range parts
                    col_c = str(row[2]).strip() if pd.notna(row[2]) else None
                    col_d = str(row[3]).strip() if pd.notna(row[3]) else None
                    col_e = str(row[4]).strip() if pd.notna(row[4]) else None
                    
                    # Build the range/value string based on what's present
                    if current_header.endswith(':'):
                        # Header ends with colon, append directly
                        if 'depth' in current_header.lower():
                            # This is a depth range (e.g., "depth to invert:")
                            if col_c and col_e:
                                if col_d == '-':
                                    # Range format: "ne - 0.5" or "0.5 - 0.75"
                                    range_str = f"{col_c} - {col_e}"
                                else:
                                    # Just use what's there
                                    range_str = f"{col_c} {col_d} {col_e}".strip()
                            elif col_c:
                                range_str = col_c
                            else:
                                range_str = col_e if col_e else ""
                            
                            # Add 'm' suffix for depth measurements
                            description = f"{current_header} {range_str} m".strip()
                        else:
                            # Generic case
                            value_str = ' '.join(filter(None, [col_c, col_d, col_e]))
                            description = f"{current_header} {value_str}".strip()
                    else:
                        # Header doesn't end with colon, use semicolon separator
                        value_str = ' '.join(filter(None, [col_c, col_d, col_e]))
                        description = f"{current_header}; {value_str}".strip()
                elif pd.notna(row[1]):
                    # Has its own description in column B
                    description = self.extract_description(row)
                else:
                    # No description available - skip
                    continue
            
            elif 309 <= row_idx <= 345:  # Section 2: Normal pattern
                # Normal items with description in column B
                if pd.notna(row[1]):
                    description = self.extract_description(row)
                else:
                    continue
            
            elif 347 <= row_idx <= 687:  # Section 3: Header + column C
                # Combine header with column C value
                if current_header and pd.notna(row[2]):  # Column C has the value
                    col_c = str(row[2]).strip()
                    if current_header.endswith(':'):
                        description = f"{current_header} {col_c}".strip()
                    else:
                        description = f"{current_header}; {col_c}".strip()
                elif pd.notna(row[1]):
                    # Has its own description in column B
                    description = self.extract_description(row)
                else:
                    continue
            
            elif 689 <= row_idx <= 806:  # Section 4: Normal pattern with bold subcategories
                # Check if description is in column B or C
                if pd.notna(row[1]):
                    description = self.extract_description(row)
                elif pd.notna(row[2]):
                    # Description is in column C for this section
                    description = str(row[2]).strip()
                else:
                    continue
            else:
                # Default case - try to get description from column B
                if pd.notna(row[1]):
                    description = self.extract_description(row)
                else:
                    continue
            
            # Skip if no valid description
            if not description or len(description) < 5:
                continue
            
            # Extract rate from column I (index 8)
            rate = None
            rate_cell_ref = None
            rate_value = None
            
            if len(row) > 8 and pd.notna(row[8]):
                try:
                    value = float(str(row[8]).replace(',', '').replace('£', ''))
                    if value > 0:
                        rate = value
                        rate_value = value
                        rate_cell_ref = f"Services!{self.get_cell_reference(row_idx, 8)}"
                except:
                    pass
            
            # Get unit
            unit = self.extract_unit(row)
            
            # Determine categories
            # Use current_subcategory for sections 2 and 4, otherwise determine from description
            if (309 <= row_idx <= 345) or (689 <= row_idx <= 806):
                subcategory = current_subcategory
            else:
                subcategory = self.determine_subcategory(description)
            
            work_type = self.determine_work_type(description, subcategory)
            
            # Generate keywords
            keywords = self.generate_keywords(description, subcategory)
            
            # Get cell references
            excel_ref = f"Services!{self.get_cell_reference(row_idx, 0)}"
            
            # Create item
            item = {
                'id': f"SV{current_id:04d}",
                'code': code,  # Use actual code from Excel
                'original_code': code,
                'description': description,
                'unit': unit,
                'category': 'Services',
                'subcategory': subcategory,
                'work_type': work_type,
                'rate': rate,
                'cellRate_reference': rate_cell_ref,
                'cellRate_rate': rate_value,
                'excelCellReference': excel_ref,
                'sourceSheetName': self.sheet_name,
                'keywords': keywords
            }
            
            items.append(item)
            current_id += 1
        
        self.extracted_items = items
        print(f"Extracted {len(items)} valid items from {self.sheet_name}")
        return items
    
    def save_output(self, output_prefix='services'):
        """Save extracted data"""
        if not self.extracted_items:
            print("No items to save")
            return
        
        # Save JSON
        json_file = f"{output_prefix}_extracted.json"
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(self.extracted_items, f, indent=2, ensure_ascii=False)
        print(f"Saved JSON: {json_file}")
        
        # Save CSV
        df = pd.DataFrame(self.extracted_items)
        df['keywords'] = df['keywords'].apply(lambda x: '|'.join(x) if x else '')
        csv_file = f"{output_prefix}_extracted.csv"
        df.to_csv(csv_file, index=False)
        print(f"Saved CSV: {csv_file}")
        
        return json_file, csv_file

def main():
    print("="*60)
    print("SERVICES SHEET EXTRACTION")
    print("="*60)
    
    extractor = ServicesExtractor()
    items = extractor.extract_items()
    
    if items:
        # Show sample
        print("\nSample extracted items:")
        for item in items[:3]:
            print(f"\nID: {item['id']}")
            print(f"  Code: {item['code']}")
            print(f"  Description: {item['description'][:60]}...")
            print(f"  Unit: {item['unit']}")
            print(f"  Subcategory: {item['subcategory']}")
            print(f"  Rate: {item['rate']}")
            print(f"  Cell Ref: {item['cellRate_reference']}")
            print(f"  Keywords: {', '.join(item['keywords'][:3])}")
        
        extractor.save_output()
        
        # Statistics
        print("\n" + "="*60)
        print("EXTRACTION STATISTICS")
        print("="*60)
        print(f"Total items: {len(items)}")
        print(f"Items with rates: {sum(1 for i in items if i['rate'])}")
        print(f"Items with cell references: {sum(1 for i in items if i['cellRate_reference'])}")
        
        # Subcategory distribution
        subcats = {}
        for item in items:
            subcat = item['subcategory']
            subcats[subcat] = subcats.get(subcat, 0) + 1
        
        print("\nSubcategory distribution:")
        for subcat, count in sorted(subcats.items(), key=lambda x: x[1], reverse=True)[:5]:
            print(f"  {subcat}: {count}")

if __name__ == "__main__":
    main()