"""
Enhanced MJD Pricelist Extraction Script with OpenAI Quality Improvement
Maintains original sheet data while improving quality and coherence
"""

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import numpy as np
import json
import csv
import re
from typing import Dict, List, Optional, Tuple, Any
from dataclasses import dataclass, asdict
from pathlib import Path
import openai
import os
from datetime import datetime
import time
import hashlib

# OpenAI Configuration
OPENAI_API_KEY = os.getenv('OPENAI_API_KEY', '')  # Set your API key as environment variable

@dataclass
class CellRate:
    """Cell reference for rate updates"""
    reference: str
    sheetName: str
    rate: float

@dataclass
class PriceItem:
    """Schema for price list item"""
    id: str
    code: Optional[str] = None
    ref: Optional[str] = None
    description: Optional[str] = None
    original_description: Optional[str] = None  # Keep original for reference
    keywords: Optional[List[str]] = None
    unit: Optional[str] = None
    category: Optional[str] = None
    subCategory: Optional[str] = None
    cellRates: Optional[Dict[str, CellRate]] = None
    patterns: Optional[List[Dict]] = None
    quality_score: Optional[float] = None  # Track quality improvements

class OpenAIEnhancer:
    """Handles OpenAI API calls for data quality improvement"""
    
    def __init__(self, api_key: str):
        self.api_key = api_key
        if api_key:
            self.client = openai.OpenAI(api_key=api_key)
        else:
            self.client = None
        self.cache = {}  # Cache responses to avoid duplicate API calls
    
    def enhance_batch_items(self, items: List[Dict], sheet_name: str) -> List[Dict]:
        """Enhance a batch of items using OpenAI"""
        if not self.client:
            return items
        
        # Process in batches of 10 to optimize API usage
        batch_size = 10
        enhanced_items = []
        
        for i in range(0, len(items), batch_size):
            batch = items[i:i + batch_size]
            enhanced_batch = self._process_batch(batch, sheet_name)
            enhanced_items.extend(enhanced_batch)
            
            # Rate limiting
            time.sleep(0.5)
        
        return enhanced_items
    
    def _process_batch(self, batch: List[Dict], sheet_name: str) -> List[Dict]:
        """Process a batch of items with OpenAI"""
        try:
            # Prepare batch data for API
            batch_text = json.dumps([{
                'index': idx,
                'description': item.get('description', ''),
                'unit': item.get('unit', ''),
                'subcategory': item.get('subCategory', ''),
                'rate': item.get('rate', '')
            } for idx, item in enumerate(batch)], indent=2)
            
            prompt = f"""You are a construction cost estimator reviewing extracted pricelist items from a spreadsheet.
            
Sheet Name: {sheet_name}
Items to review:
{batch_text}

For each item, improve the data quality while PRESERVING the original meaning and intent:

1. Description: 
   - If truncated or fragmented, complete it based on construction context
   - Fix typos and formatting issues
   - Make it clear and professional
   - DO NOT change the fundamental meaning
   - If it references specific products/standards (BS, EN, etc.), keep them

2. Unit:
   - Standardize units (m², m³, nr, m, kg, tonnes, hours, days, weeks)
   - If missing but obvious from description, suggest appropriate unit
   
3. SubCategory:
   - Suggest proper subcategory based on the description
   - Common categories for {sheet_name}:
     * Groundworks: Site Clearance, Excavation, Earthworks Support, Disposal, Filling
     * RC Works: In-situ Concrete, Formwork, Reinforcement, Precast, Post-tensioning
     * Drainage: Below Ground Drainage, Manholes, Pipework, Gullies, Connections
     * External Works: Paving, Kerbs, Fencing, Landscaping, Street Furniture
     * Services: Electrical, Mechanical, Plumbing, HVAC, Fire Systems
     * Prelims: Management, Site Setup, Temporary Works, Welfare, Plant

4. Keywords:
   - Generate 3-5 relevant search keywords for matching
   
Return a JSON array with the enhanced items in this exact format:
[
  {{
    "index": 0,
    "description": "improved description",
    "unit": "standardized unit",
    "subCategory": "appropriate subcategory",
    "keywords": ["keyword1", "keyword2", "keyword3"],
    "quality_notes": "brief note on what was improved"
  }}
]

IMPORTANT: 
- Maintain the same order (use index)
- Keep descriptions concise but complete
- Use standard construction terminology
- Return valid JSON only"""

            response = self.client.chat.completions.create(
                model="gpt-4-turbo-preview",
                messages=[
                    {"role": "system", "content": "You are an expert construction cost estimator with deep knowledge of BOQ items and construction terminology."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.3,
                max_tokens=2000,
                response_format={"type": "json_object"}
            )
            
            # Parse response
            result_text = response.choices[0].message.content.strip()
            
            # Handle potential JSON wrapped in markdown
            if '```json' in result_text:
                result_text = result_text.split('```json')[1].split('```')[0].strip()
            elif '```' in result_text:
                result_text = result_text.split('```')[1].split('```')[0].strip()
            
            # Ensure we have a proper JSON object
            if not result_text.startswith('[') and not result_text.startswith('{'):
                # Try to find JSON array in the response
                import re
                json_match = re.search(r'\[[\s\S]*\]', result_text)
                if json_match:
                    result_text = json_match.group()
            
            # Parse JSON response
            if result_text.startswith('{') and 'items' in result_text:
                enhanced_data = json.loads(result_text).get('items', [])
            else:
                enhanced_data = json.loads(result_text)
            
            # Merge enhanced data back into original items
            enhanced_batch = []
            for item in batch:
                # Find corresponding enhanced item
                enhanced_item = item.copy()
                
                for enhanced in enhanced_data:
                    if enhanced.get('index') == batch.index(item):
                        # Preserve original description
                        enhanced_item['original_description'] = item.get('description', '')
                        
                        # Apply improvements
                        if enhanced.get('description'):
                            enhanced_item['description'] = enhanced['description']
                        
                        if enhanced.get('unit'):
                            enhanced_item['unit'] = enhanced['unit']
                        
                        if enhanced.get('subCategory'):
                            enhanced_item['subCategory'] = enhanced['subCategory']
                        
                        if enhanced.get('keywords'):
                            enhanced_item['keywords'] = enhanced['keywords']
                        
                        # Add quality score based on improvements
                        enhanced_item['quality_score'] = 1.0 if enhanced.get('quality_notes') else 0.5
                        
                        break
                
                enhanced_batch.append(enhanced_item)
            
            return enhanced_batch
            
        except Exception as e:
            print(f"Error enhancing batch with OpenAI: {e}")
            # Return original items if enhancement fails
            return batch
    
    def validate_unit(self, unit: str, description: str) -> str:
        """Validate and standardize units using OpenAI"""
        if not self.client or not unit:
            return unit
        
        # Cache key
        cache_key = f"unit_{unit}_{description[:50]}"
        if cache_key in self.cache:
            return self.cache[cache_key]
        
        try:
            prompt = f"""Standardize this construction unit of measurement:
            Current unit: "{unit}"
            Item description: "{description}"
            
            Return ONLY the standardized unit (m, m², m³, nr, kg, tonnes, hours, days, weeks, sum, item).
            If the unit is already correct, return it as-is."""
            
            response = self.client.chat.completions.create(
                model="gpt-3.5-turbo",
                messages=[{"role": "user", "content": prompt}],
                temperature=0,
                max_tokens=10
            )
            
            standardized = response.choices[0].message.content.strip()
            self.cache[cache_key] = standardized
            return standardized
            
        except:
            return unit

class SheetExtractor:
    """Base class for sheet-specific extraction logic"""
    
    def __init__(self, sheet_name: str, worksheet, enhancer: Optional[OpenAIEnhancer] = None):
        self.sheet_name = sheet_name
        self.worksheet = worksheet
        self.enhancer = enhancer
        self.items = []
        
    def get_cell_reference(self, row: int, col: int) -> str:
        """Get Excel-style cell reference (e.g., A1, B2)"""
        return f"{get_column_letter(col)}{row}"
    
    def extract_header_info(self) -> Dict:
        """Extract header and subcategory information"""
        headers = {}
        for row in range(1, min(20, self.worksheet.max_row + 1)):
            for col in range(1, min(10, self.worksheet.max_column + 1)):
                cell_value = self.worksheet.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str):
                    if any(keyword in str(cell_value).lower() for keyword in ['schedule', 'works', 'section']):
                        headers['main_header'] = cell_value
                        headers['header_row'] = row
                        break
            if headers:
                break
        return headers
    
    def clean_value(self, value: Any) -> Any:
        """Clean cell values"""
        if value is None:
            return None
        if isinstance(value, str):
            # Remove excessive whitespace
            cleaned = ' '.join(value.split())
            # Remove non-printable characters
            cleaned = ''.join(char for char in cleaned if char.isprintable())
            return cleaned
        return value
    
    def is_valid_rate(self, value: Any) -> bool:
        """Check if value is a valid rate"""
        if value is None:
            return False
        try:
            rate = float(value)
            return 0 < rate < 1000000  # Reasonable rate range
        except:
            return False
    
    def is_valid_description(self, desc: str) -> bool:
        """Check if description is valid"""
        if not desc:
            return False
        
        # Skip if too short
        if len(desc) < 3:
            return False
        
        # Skip if it's just numbers or special characters
        if not any(c.isalpha() for c in desc):
            return False
        
        # Skip common non-item texts
        skip_terms = ['total', 'subtotal', 'carried forward', 'brought forward', 
                     'continued', 'see over', 'blank', 'n/a', 'nil', '-']
        
        desc_lower = desc.lower()
        if any(term in desc_lower for term in skip_terms):
            return False
        
        return True
    
    def extract_items(self) -> List[PriceItem]:
        """Main extraction method - must be implemented by subclasses"""
        raise NotImplementedError("Each sheet extractor must implement extract_items")
    
    def post_process_items(self, items: List[PriceItem]) -> List[PriceItem]:
        """Post-process items with OpenAI enhancement"""
        if not self.enhancer or not items:
            return items
        
        print(f"  Enhancing {len(items)} items with OpenAI...")
        
        # Convert items to dict for enhancement
        items_dict = []
        for item in items:
            item_dict = {
                'description': item.description,
                'unit': item.unit,
                'subCategory': item.subCategory,
                'rate': item.cellRates['cellRate1'].rate if item.cellRates and 'cellRate1' in item.cellRates else None
            }
            items_dict.append(item_dict)
        
        # Enhance with OpenAI
        enhanced_dicts = self.enhancer.enhance_batch_items(items_dict, self.sheet_name)
        
        # Apply enhancements back to items
        for i, item in enumerate(items):
            if i < len(enhanced_dicts):
                enhanced = enhanced_dicts[i]
                
                # Store original description
                item.original_description = item.description
                
                # Apply enhancements
                if enhanced.get('description'):
                    item.description = enhanced['description']
                
                if enhanced.get('unit'):
                    item.unit = enhanced['unit']
                
                if enhanced.get('subCategory'):
                    item.subCategory = enhanced['subCategory']
                
                if enhanced.get('keywords'):
                    item.keywords = enhanced['keywords']
                
                if enhanced.get('quality_score'):
                    item.quality_score = enhanced['quality_score']
        
        return items

class GroundworksExtractor(SheetExtractor):
    """Extractor for Groundworks sheet with enhanced quality"""
    
    def extract_items(self) -> List[PriceItem]:
        items = []
        current_subcategory = None
        item_counter = 1
        
        # Find data start row
        data_start_row = 10
        for row in range(1, min(30, self.worksheet.max_row + 1)):
            cell = self.worksheet.cell(row=row, column=2)
            if cell.value and 'Description' in str(cell.value):
                data_start_row = row + 2
                break
        
        # Track potential subcategory headers
        potential_subcategories = []
        
        # Process rows
        for row in range(data_start_row, self.worksheet.max_row + 1):
            # Check for subcategory headers
            cell_b = self.worksheet.cell(row=row, column=2)
            
            # Check if this is a subcategory header (bold, merged, or all caps)
            if cell_b.value:
                cell_value = self.clean_value(cell_b.value)
                
                # Check for subcategory indicators
                is_subcategory = False
                
                # Check if bold
                if cell_b.font and cell_b.font.bold:
                    is_subcategory = True
                
                # Check if all caps and short
                if cell_value and cell_value.isupper() and len(cell_value) < 50:
                    is_subcategory = True
                
                # Check for known groundworks subcategories
                subcategory_keywords = ['excavat', 'disposal', 'filling', 'earthwork', 
                                       'foundation', 'drainage', 'concrete', 'piling']
                if any(keyword in cell_value.lower() for keyword in subcategory_keywords):
                    is_subcategory = True
                
                if is_subcategory:
                    current_subcategory = cell_value
                    potential_subcategories.append(cell_value)
                    continue
            
            # Extract item data
            description = self.clean_value(cell_b.value)
            
            # Validate description
            if not self.is_valid_description(description):
                continue
            
            # Get the item code if present (often in column A)
            item_code = self.clean_value(self.worksheet.cell(row=row, column=1).value)
            
            # Get quantity, unit, rate columns
            quantity = self.worksheet.cell(row=row, column=4).value
            unit = self.clean_value(self.worksheet.cell(row=row, column=5).value)
            rate = self.worksheet.cell(row=row, column=6).value
            
            # Try to find rate in adjacent columns if not in column 6
            if not self.is_valid_rate(rate):
                for col in [7, 8, 9]:
                    alt_rate = self.worksheet.cell(row=row, column=col).value
                    if self.is_valid_rate(alt_rate):
                        rate = alt_rate
                        break
            
            # Only process rows with valid data
            if description and (rate or quantity):
                # Use original code if present, otherwise generate
                if item_code and str(item_code).strip():
                    code = str(item_code).strip()
                else:
                    code = f"GW{item_counter:04d}"
                
                # Generate ID using sheet name and code
                item_id = f"{self.sheet_name}_{code}"
                
                # Get cell references for rates
                cell_rates = {}
                if self.is_valid_rate(rate):
                    cell_rates['cellRate1'] = CellRate(
                        reference=self.get_cell_reference(row, 6),
                        sheetName=self.sheet_name,
                        rate=float(rate) if rate else 0.0
                    )
                
                # Check for additional rate columns
                for col_offset, rate_name in enumerate(['cellRate2', 'cellRate3', 'cellRate4'], start=1):
                    additional_rate = self.worksheet.cell(row=row, column=6 + col_offset).value
                    if self.is_valid_rate(additional_rate):
                        cell_rates[rate_name] = CellRate(
                            reference=self.get_cell_reference(row, 6 + col_offset),
                            sheetName=self.sheet_name,
                            rate=float(additional_rate)
                        )
                
                # Infer subcategory if not set
                if not current_subcategory:
                    desc_lower = description.lower()
                    if 'excavat' in desc_lower:
                        current_subcategory = 'Excavation'
                    elif 'fill' in desc_lower:
                        current_subcategory = 'Filling'
                    elif 'disposal' in desc_lower or 'cart away' in desc_lower:
                        current_subcategory = 'Disposal'
                    elif 'clear' in desc_lower or 'demolish' in desc_lower:
                        current_subcategory = 'Site Clearance'
                
                item = PriceItem(
                    id=item_id,
                    code=code,
                    description=description,
                    unit=unit,
                    category=self.sheet_name,
                    subCategory=current_subcategory,
                    cellRates=cell_rates if cell_rates else None
                )
                
                items.append(item)
                item_counter += 1
        
        # Post-process with OpenAI if available
        return self.post_process_items(items)

class RCWorksExtractor(SheetExtractor):
    """Extractor for RC Works sheet with enhanced quality"""
    
    def extract_items(self) -> List[PriceItem]:
        items = []
        current_subcategory = None
        item_counter = 1
        
        # RC Works specific structure
        data_start_row = 10
        for row in range(1, min(30, self.worksheet.max_row + 1)):
            cell = self.worksheet.cell(row=row, column=2)
            if cell.value and 'Description' in str(cell.value):
                data_start_row = row + 2
                break
        
        for row in range(data_start_row, self.worksheet.max_row + 1):
            # Column B for descriptions
            cell_b = self.worksheet.cell(row=row, column=2)
            
            # Check for subcategory headers
            if cell_b.value:
                value_str = self.clean_value(cell_b.value)
                
                # RC specific subcategories
                rc_categories = {
                    'in-situ concrete': 'In-situ Concrete',
                    'in situ concrete': 'In-situ Concrete',
                    'formwork': 'Formwork',
                    'reinforcement': 'Reinforcement',
                    'rebar': 'Reinforcement',
                    'precast': 'Precast Concrete',
                    'post-tension': 'Post-tensioning',
                    'prestress': 'Post-tensioning',
                    'sundries': 'Concrete Sundries',
                    'accessories': 'Concrete Accessories'
                }
                
                value_lower = value_str.lower() if value_str else ''
                for key, category in rc_categories.items():
                    if key in value_lower:
                        current_subcategory = category
                        continue
            
            description = self.clean_value(cell_b.value)
            
            if not self.is_valid_description(description):
                continue
            
            # Get item code
            item_code = self.clean_value(self.worksheet.cell(row=row, column=1).value)
            
            # Get data from standard columns
            quantity = self.worksheet.cell(row=row, column=4).value
            unit = self.clean_value(self.worksheet.cell(row=row, column=5).value)
            rate = self.worksheet.cell(row=row, column=6).value
            
            if description and (rate or quantity):
                # Use original code or generate
                if item_code and str(item_code).strip():
                    code = str(item_code).strip()
                else:
                    code = f"RC{item_counter:04d}"
                
                item_id = f"{self.sheet_name}_{code}"
                
                cell_rates = {}
                if self.is_valid_rate(rate):
                    cell_rates['cellRate1'] = CellRate(
                        reference=self.get_cell_reference(row, 6),
                        sheetName=self.sheet_name,
                        rate=float(rate) if rate else 0.0
                    )
                
                # Auto-detect subcategory from description if needed
                if not current_subcategory:
                    desc_lower = description.lower()
                    if 'concrete' in desc_lower and 'formwork' not in desc_lower:
                        current_subcategory = 'In-situ Concrete'
                    elif 'formwork' in desc_lower or 'shutter' in desc_lower:
                        current_subcategory = 'Formwork'
                    elif 'reinforc' in desc_lower or 'rebar' in desc_lower or 'mesh' in desc_lower:
                        current_subcategory = 'Reinforcement'
                
                item = PriceItem(
                    id=item_id,
                    code=code,
                    description=description,
                    unit=unit,
                    category=self.sheet_name,
                    subCategory=current_subcategory,
                    cellRates=cell_rates if cell_rates else None
                )
                
                items.append(item)
                item_counter += 1
        
        return self.post_process_items(items)

class DrainageExtractor(SheetExtractor):
    """Extractor for Drainage sheet with enhanced quality"""
    
    def extract_items(self) -> List[PriceItem]:
        items = []
        current_subcategory = None
        item_counter = 1
        
        # Find actual data start
        data_start_row = 15
        for row in range(1, min(50, self.worksheet.max_row + 1)):
            cell = self.worksheet.cell(row=row, column=2)
            if cell.value and any(term in str(cell.value).lower() for term in ['pipe', 'excavat', 'manhole', 'drain']):
                data_start_row = max(row - 2, 1)  # Start a bit before first item
                break
        
        for row in range(data_start_row, self.worksheet.max_row + 1):
            # Get description from multiple possible columns
            description = None
            desc_col = None
            
            for col in [2, 3, 4]:
                cell_value = self.worksheet.cell(row=row, column=col).value
                if cell_value:
                    cleaned = self.clean_value(cell_value)
                    if cleaned and len(cleaned) > 5 and any(c.isalpha() for c in cleaned):
                        description = cleaned
                        desc_col = col
                        break
            
            if not description:
                continue
            
            # Check for subcategory markers
            desc_lower = description.lower()
            drainage_categories = {
                'below ground': 'Below Ground Drainage',
                'above ground': 'Above Ground Drainage',
                'manhole': 'Manholes & Chambers',
                'chamber': 'Manholes & Chambers',
                'gully': 'Gullies & Gratings',
                'grating': 'Gullies & Gratings',
                'pipe': 'Pipework',
                'excavat': 'Excavation for Drainage',
                'bedding': 'Pipe Bedding',
                'connection': 'Connections',
                'testing': 'Testing & Commissioning'
            }
            
            # Check if this is a subcategory header
            is_header = False
            for key, cat in drainage_categories.items():
                if key in desc_lower and len(description) < 50:
                    current_subcategory = cat
                    is_header = True
                    break
            
            if is_header:
                continue
            
            # Skip non-items
            if not self.is_valid_description(description):
                continue
            
            # Get item code
            item_code = self.clean_value(self.worksheet.cell(row=row, column=1).value)
            
            # Find rate and unit columns
            rate = None
            rate_col = None
            unit = None
            
            # Look for rate in standard positions
            for col in [6, 7, 8, 9, 10]:
                cell_value = self.worksheet.cell(row=row, column=col).value
                if self.is_valid_rate(cell_value):
                    rate = cell_value
                    rate_col = col
                    # Unit is usually before rate
                    unit = self.clean_value(self.worksheet.cell(row=row, column=col - 1).value)
                    break
            
            if description and rate:
                # Use original code or generate
                if item_code and str(item_code).strip():
                    code = str(item_code).strip()
                else:
                    code = f"DR{item_counter:04d}"
                
                item_id = f"{self.sheet_name}_{code}"
                
                cell_rates = {}
                if rate and rate_col:
                    cell_rates['cellRate1'] = CellRate(
                        reference=self.get_cell_reference(row, rate_col),
                        sheetName=self.sheet_name,
                        rate=float(rate)
                    )
                
                # Auto-detect subcategory if needed
                if not current_subcategory:
                    desc_lower = description.lower()
                    for key, cat in drainage_categories.items():
                        if key in desc_lower:
                            current_subcategory = cat
                            break
                    
                    if not current_subcategory:
                        current_subcategory = 'General Drainage'
                
                item = PriceItem(
                    id=item_id,
                    code=code,
                    description=description,
                    unit=unit,
                    category=self.sheet_name,
                    subCategory=current_subcategory,
                    cellRates=cell_rates if cell_rates else None
                )
                
                items.append(item)
                item_counter += 1
        
        return self.post_process_items(items)

class EnhancedPricelistExtractor:
    """Main extractor with OpenAI enhancement"""
    
    SHEET_EXTRACTORS = {
        'Groundworks': GroundworksExtractor,
        'RC works': RCWorksExtractor,
        'Drainage': DrainageExtractor,
        # Add more as needed
    }
    
    SKIP_SHEETS = ['Summary', 'Set factors & prices', 'Tender Summary', 'Budget Costings']
    
    def __init__(self, file_path: str, use_openai: bool = True):
        self.file_path = file_path
        self.use_openai = use_openai
        self.workbook = None
        self.all_items = []
        
        # Initialize OpenAI enhancer
        if use_openai and OPENAI_API_KEY:
            self.enhancer = OpenAIEnhancer(OPENAI_API_KEY)
        else:
            self.enhancer = None
            if use_openai:
                print("Warning: OpenAI API key not found. Enhancement disabled.")
    
    def extract_all_sheets(self) -> List[PriceItem]:
        """Extract items from all sheets with quality enhancement"""
        self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
        
        for sheet_name in self.workbook.sheetnames:
            if sheet_name in self.SKIP_SHEETS:
                print(f"Skipping sheet: {sheet_name}")
                continue
            
            print(f"Processing sheet: {sheet_name}")
            worksheet = self.workbook[sheet_name]
            
            # Select appropriate extractor or use base class
            if sheet_name in self.SHEET_EXTRACTORS:
                extractor_class = self.SHEET_EXTRACTORS[sheet_name]
            else:
                # Use base extractor with generic logic
                extractor_class = SheetExtractor
            
            # Create extractor with enhancer
            if extractor_class == SheetExtractor:
                # For base class, we need to implement a generic extraction
                extractor = GenericEnhancedExtractor(sheet_name, worksheet, self.enhancer)
            else:
                extractor = extractor_class(sheet_name, worksheet, self.enhancer)
            
            # Extract items
            items = extractor.extract_items()
            
            self.all_items.extend(items)
            print(f"  Extracted and enhanced {len(items)} items from {sheet_name}")
        
        self.workbook.close()
        
        # Final quality report
        self.print_quality_report()
        
        return self.all_items
    
    def print_quality_report(self):
        """Print quality improvement report"""
        if not self.all_items:
            return
        
        enhanced_count = sum(1 for item in self.all_items if hasattr(item, 'original_description') and item.original_description)
        with_keywords = sum(1 for item in self.all_items if item.keywords)
        with_subcategory = sum(1 for item in self.all_items if item.subCategory)
        
        print("\n" + "="*50)
        print("Quality Enhancement Report")
        print("="*50)
        print(f"Total items processed: {len(self.all_items)}")
        print(f"Items enhanced: {enhanced_count}")
        print(f"Items with keywords: {with_keywords}")
        print(f"Items with subcategories: {with_subcategory}")
        
        if enhanced_count > 0:
            print("\nSample improvements:")
            samples = [item for item in self.all_items if hasattr(item, 'original_description') and item.original_description][:3]
            for item in samples:
                print(f"\n  Original: {item.original_description[:60]}...")
                print(f"  Enhanced: {item.description[:60]}...")
                if item.keywords:
                    print(f"  Keywords: {', '.join(item.keywords[:5])}")
    
    def export_to_csv(self, output_file: str = "pricelist_enhanced.csv"):
        """Export enhanced items to CSV"""
        if not self.all_items:
            print("No items to export")
            return
        
        csv_data = []
        for item in self.all_items:
            row = {
                'id': item.id,
                'code': item.code or '',
                'ref': item.ref or '',
                'description': item.description or '',
                'original_description': getattr(item, 'original_description', ''),
                'unit': item.unit or '',
                'category': item.category or '',
                'subCategory': item.subCategory or '',
                'keywords': '|'.join(item.keywords) if item.keywords else '',
                'quality_score': getattr(item, 'quality_score', 0.0)
            }
            
            # Add cell rates
            if item.cellRates:
                for i, (rate_key, rate_obj) in enumerate(item.cellRates.items(), 1):
                    row[f'cellRate{i}_reference'] = rate_obj.reference
                    row[f'cellRate{i}_sheetName'] = rate_obj.sheetName
                    row[f'cellRate{i}_rate'] = rate_obj.rate
            
            csv_data.append(row)
        
        # Write to CSV
        if csv_data:
            fieldnames = list(csv_data[0].keys())
            
            with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(csv_data)
            
            print(f"\nExported {len(csv_data)} enhanced items to {output_file}")

class GenericEnhancedExtractor(SheetExtractor):
    """Generic extractor with OpenAI enhancement for unspecified sheets"""
    
    def extract_items(self) -> List[PriceItem]:
        items = []
        current_subcategory = None
        item_counter = 1
        
        # Auto-detect column positions
        desc_col = 2
        rate_col = 6
        unit_col = 5
        code_col = 1
        
        # Find headers
        for row in range(1, min(20, self.worksheet.max_row + 1)):
            for col in range(1, min(15, self.worksheet.max_column + 1)):
                cell_value = str(self.worksheet.cell(row=row, column=col).value or '').lower()
                if 'description' in cell_value:
                    desc_col = col
                elif 'rate' in cell_value:
                    rate_col = col
                elif 'unit' in cell_value:
                    unit_col = col
                elif 'code' in cell_value or 'ref' in cell_value:
                    code_col = col
        
        # Start extraction
        data_start_row = 10
        
        for row in range(data_start_row, self.worksheet.max_row + 1):
            description = self.clean_value(self.worksheet.cell(row=row, column=desc_col).value)
            
            if not self.is_valid_description(description):
                continue
            
            item_code = self.clean_value(self.worksheet.cell(row=row, column=code_col).value)
            unit = self.clean_value(self.worksheet.cell(row=row, column=unit_col).value)
            rate = self.worksheet.cell(row=row, column=rate_col).value
            
            if description and self.is_valid_rate(rate):
                # Generate code
                if item_code and str(item_code).strip():
                    code = str(item_code).strip()
                else:
                    prefix = ''.join([c for c in self.sheet_name[:2].upper() if c.isalpha()])[:2]
                    if not prefix:
                        prefix = 'GN'
                    code = f"{prefix}{item_counter:04d}"
                
                item_id = f"{self.sheet_name}_{code}"
                
                cell_rates = {
                    'cellRate1': CellRate(
                        reference=self.get_cell_reference(row, rate_col),
                        sheetName=self.sheet_name,
                        rate=float(rate)
                    )
                }
                
                item = PriceItem(
                    id=item_id,
                    code=code,
                    description=description,
                    unit=unit,
                    category=self.sheet_name,
                    subCategory=current_subcategory,
                    cellRates=cell_rates
                )
                
                items.append(item)
                item_counter += 1
        
        return self.post_process_items(items)

def main():
    """Main execution function"""
    file_path = r"C:\Users\abaza\pricelist extraction\MJD-PRICELIST.xlsx"
    
    if not Path(file_path).exists():
        print(f"Error: File not found at {file_path}")
        return
    
    print("="*50)
    print("Enhanced MJD Pricelist Extraction Tool")
    print("="*50)
    
    # Check for OpenAI API key
    if not OPENAI_API_KEY:
        print("\n⚠️  WARNING: No OpenAI API key found!")
        print("Set OPENAI_API_KEY environment variable for quality enhancement.")
        print("Example: set OPENAI_API_KEY=sk-...")
        response = input("\nContinue without enhancement? (y/n): ")
        if response.lower() != 'y':
            return
        use_openai = False
    else:
        print("✓ OpenAI API key found - Enhancement enabled")
        use_openai = True
    
    # Create extractor
    extractor = EnhancedPricelistExtractor(file_path, use_openai=use_openai)
    
    # Extract all sheets
    print("\nStarting enhanced extraction...")
    items = extractor.extract_all_sheets()
    
    print(f"\nTotal items extracted: {len(items)}")
    
    # Export results
    print("\nExporting enhanced results...")
    extractor.export_to_csv("pricelist_enhanced.csv")
    
    print("\n✓ Enhanced extraction complete!")

if __name__ == "__main__":
    main()